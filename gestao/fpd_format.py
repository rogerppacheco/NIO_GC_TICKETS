from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import RelatorioFPD

FAIXAS_ORDEM = [
    "10 a 15 Dias",
    "15 a 30 Dias",
    "30 a 45 Dias",
    "45 a 55 Dias",
    "55 a 60 Dias",
    ">= a 61 Dias",
]

FAIXA_ROTULO = {
    "10 a 15 Dias": "10 a 15 Dias:",
    "15 a 30 Dias": "15 a 30 Dias:",
    "30 a 45 Dias": "30 a 45 Dias:",
    "45 a 55 Dias": "45 a 55 Dias:",
    "55 a 60 Dias": "55 a 60 Dias:",
    ">= a 61 Dias": ">= a 61 Dias:",
}

FAIXA_CORES = {
    "10 a 15 Dias": ("FFFF99", "000000"),
    "15 a 30 Dias": ("FFFF00", "000000"),
    "30 a 45 Dias": ("FFCCCC", "000000"),
    "45 a 55 Dias": ("FF6666", "000000"),
    "55 a 60 Dias": ("FF0000", "FFFFFF"),
    ">= a 61 Dias": ("8B0000", "FFFFFF"),
}


def _serializar_celula(valor: Any) -> Any:
    if valor is None:
        return None
    try:
        if pd.isna(valor):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(valor, (datetime, date)):
        return valor.isoformat(sep=" ", timespec="seconds") if isinstance(valor, datetime) else valor.isoformat()
    if isinstance(valor, float):
        if valor.is_integer():
            return int(valor)
        return valor
    return valor


def dataframe_para_base(df: pd.DataFrame) -> list[dict]:
    if df.empty:
        return []
    registros = []
    for row in df.to_dict(orient="records"):
        registros.append({str(k): _serializar_celula(v) for k, v in row.items()})
    return registros


def base_para_dataframe(detalhes: dict) -> pd.DataFrame:
    colunas = detalhes.get("base_colunas") or []
    linhas = detalhes.get("base") or []
    if not linhas:
        return pd.DataFrame(columns=colunas)
    df = pd.DataFrame(linhas)
    if colunas:
        for col in colunas:
            if col not in df.columns:
                df[col] = None
        extras = [c for c in df.columns if c not in colunas]
        df = df[colunas + extras]
    return df


def mes_para_yyyymm(valor) -> str:
    if valor is None:
        return ""
    texto = str(valor).strip()
    if texto.endswith(".0") and texto[:-2].isdigit():
        texto = texto[:-2]
    if texto.isdigit() and len(texto) == 6:
        return texto
    if "/" in texto:
        try:
            a, b = texto.split("/", 1)
            if len(a) == 4:
                return f"{int(a):04d}{int(b):02d}"
            return f"{int(b):04d}{int(a):02d}"
        except ValueError:
            return texto
    return texto


def _meses_ordenados(detalhes: dict) -> list[dict]:
    meses = list((detalhes or {}).get("meses") or [])
    return sorted(meses, key=lambda m: mes_para_yyyymm(m.get("mes_yyyymm") or m.get("mes")))


def _intervalo_meses(meses: list[dict]) -> str:
    chaves = [mes_para_yyyymm(m.get("mes_yyyymm") or m.get("mes")) for m in meses]
    chaves = [c for c in chaves if c]
    if not chaves:
        return ""
    if len(chaves) == 1:
        return chaves[0]
    return f"{chaves[0]} a {chaves[-1]}"


def _tag_arquivo(nome: str) -> str:
    limpo = re.sub(r"[^\w\-]+", "_", (nome or "PDV").strip(), flags=re.UNICODE)
    return limpo.strip("_")[:50] or "PDV"


def _codigo_rede(rel: RelatorioFPD) -> str:
    det = rel.detalhes or {}
    codigo = str(det.get("codigo_rede") or "").strip()
    if codigo.endswith(".0") and codigo[:-2].isdigit():
        codigo = codigo[:-2]
    return codigo


def _total_pagas(rel: RelatorioFPD) -> int:
    det = rel.detalhes or {}
    if det.get("total_pagas") is not None:
        return int(det["total_pagas"])
    meses = _meses_ordenados(det)
    if meses:
        return sum(int(m.get("pagas") or 0) for m in meses)
    return max(rel.total_faturas - rel.total_abertas, 0)


def _fmt_percentual_br(valor: float) -> str:
    return f"{valor:.2f}".replace(".", ",") + "%"


def _tabela_resumo(rel: RelatorioFPD) -> list[list[str]]:
    meses = _meses_ordenados(rel.detalhes or {})
    cabecalho = ["MÊS FATURA"] + [mes_para_yyyymm(m.get("mes_yyyymm") or m.get("mes")) for m in meses]
    linhas = [cabecalho]
    linhas.append(["FATURA PAGA"] + [str(int(m.get("pagas") or 0)) for m in meses])
    linhas.append(["TOTAL FATURA"] + [str(int(m.get("total") or 0)) for m in meses])
    linhas.append(
        ["% ABERTO"]
        + [
            _fmt_percentual_br(
                (int(m.get("abertas") or 0) / int(m.get("total") or 1) * 100)
                if int(m.get("total") or 0)
                else 0.0
            )
            for m in meses
        ]
    )
    for faixa in FAIXAS_ORDEM:
        linha = [FAIXA_ROTULO[faixa]]
        for mes in meses:
            faixas = mes.get("faixas") or {}
            linha.append(str(int(faixas.get(faixa) or 0)))
        linhas.append(linha)
    return linhas


def _estilo_faixa(label: str) -> tuple[str, str]:
    for chave, estilo in FAIXA_CORES.items():
        if chave in label:
            return estilo
    return ("FFFFFF", "000000")


def planilha_fpd(rel: RelatorioFPD) -> tuple[bytes, str]:
    wb = Workbook()
    ws_resumo = wb.active
    ws_resumo.title = "Planilha1"

    tabela = _tabela_resumo(rel)
    for r_idx, linha in enumerate(tabela, start=1):
        for c_idx, valor in enumerate(linha, start=1):
            cell = ws_resumo.cell(row=r_idx, column=c_idx, value=valor)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if r_idx == 1 or c_idx == 1:
                cell.font = Font(bold=True)
            if r_idx >= 5 and c_idx > 1:
                fundo, fonte = _estilo_faixa(linha[0])
                cell.fill = PatternFill("solid", fgColor=fundo)
                cell.font = Font(color=fonte, bold=True)
            if linha[0] == "% ABERTO" and c_idx > 1:
                cell.font = Font(bold=True, underline="single")

    for col in range(1, len(tabela[0]) + 1):
        ws_resumo.column_dimensions[get_column_letter(col)].width = 16

    ws_base = wb.create_sheet("BASE_PRE_FPD_ABERTO")
    df_base = base_para_dataframe(rel.detalhes or {})
    if df_base.empty:
        ws_base.append(["Sem base detalhada para este PDV."])
    else:
        ws_base.append(list(df_base.columns))
        for row in df_base.itertuples(index=False, name=None):
            ws_base.append([_serializar_celula(v) for v in row])
        for cell in ws_base[1]:
            cell.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    codigo = _codigo_rede(rel)
    sufixo = f"{codigo}-" if codigo else ""
    nome = f"FATURAS_ABERTAS_PRE-FIBRA-FPD-SPD-TPD{sufixo}{_tag_arquivo(rel.pdv_nome)}.xlsx"
    return buf.getvalue(), nome


def assunto_email_fpd(rel: RelatorioFPD) -> str:
    codigo = _codigo_rede(rel)
    pdv = (rel.pdv_nome or rel.parceiro.nome or "PDV").strip().upper()
    if codigo:
        return f"FATURAS_ABERTAS_PRÉ-FIBRA-FPD-SPD-TPD{codigo}-{pdv}"
    return f"FATURAS_ABERTAS_PRÉ-FIBRA-FPD-SPD-TPD-{pdv}"


def html_email_fpd(rel: RelatorioFPD) -> str:
    pdv = (rel.pdv_nome or rel.parceiro.nome or "PDV").strip().upper()
    meses = _meses_ordenados(rel.detalhes or {})
    intervalo = _intervalo_meses(meses)
    total = rel.total_faturas
    pagas = _total_pagas(rel)
    abertas = rel.total_abertas
    perc = rel.percentual

    linhas_html = []
    tabela = _tabela_resumo(rel)
    for r_idx, linha in enumerate(tabela):
        cells = []
        for c_idx, valor in enumerate(linha):
            estilo = ""
            if r_idx == 0 and c_idx > 0:
                estilo = ' style="font-size:13.5pt;font-weight:bold;text-decoration:underline"'
            elif linha[0] == "% ABERTO" and c_idx > 0:
                estilo = ' style="font-weight:bold;text-decoration:underline"'
            elif r_idx >= 4 and c_idx > 0 and str(valor) not in {"0", "0,00%"}:
                fundo, cor = _estilo_faixa(linha[0])
                estilo = f' style="background:#{fundo};color:#{cor};font-weight:bold"'
            cells.append(
                f'<td align="center" style="text-align:center"{estilo}>'
                f'<p align="center" style="text-align:center;margin:0">{valor}</p></td>'
            )
        linhas_html.append("<tr>" + "".join(cells) + "</tr>")

    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head><meta charset="utf-8"></head>
<body bgcolor="#EBF0EA" style="font-family:Calibri,Arial,sans-serif;color:#000">
<p><b>Bom dia, prezado parceiro!</b></p>
<p><b><span style="font-size:21pt;color:#21C002">{pdv}</span></b></p>
<p>Segue <b>Faturas abertas</b> (FPD/SPD/TPD) <b>15 a 60 dias em aberto com vencimento no meses de {intervalo}</b><br>
Faturas Abertas com vencimento menor que 61 dias.<br>
<b>{pdv}</b> - Com o total de faturas de
<b><u><span style="font-size:18pt;color:blue">{total}</span></u></b> e
<b><u><span style="font-size:18pt;color:green">{pagas}</span></u></b> Pagas, sendo com risco de FPD
<b><u><span style="font-size:18pt;color:red">{abertas}</span></u></b> com o Percentual de
<b><u><span style="font-size:18pt;color:red">{perc:.2f}%</span></u></b> Das Faturas Totais<br>
<b>Faixas e Quantidades, Faturas Abertas a tratar</b></p>
<table border="1" cellpadding="4" cellspacing="0" style="background:white;border-collapse:collapse">
{"".join(linhas_html)}
</table>
<p>Planilha detalhada em anexo.</p>
</body></html>"""


def corpo_texto_email_fpd(rel: RelatorioFPD) -> str:
    pdv = (rel.pdv_nome or rel.parceiro.nome or "PDV").strip().upper()
    intervalo = _intervalo_meses(_meses_ordenados(rel.detalhes or {}))
    return (
        f"Bom dia, prezado parceiro!\n\n"
        f"{pdv}\n\n"
        f"Segue Faturas abertas (FPD/SPD/TPD) 15 a 60 dias em aberto "
        f"com vencimento nos meses de {intervalo}.\n"
        f"Total: {rel.total_faturas} | Pagas: {_total_pagas(rel)} | "
        f"Em aberto (FPD): {rel.total_abertas} | Percentual: {rel.percentual:.2f}%\n\n"
        f"Planilha detalhada em anexo."
    )
