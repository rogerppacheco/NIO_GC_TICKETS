from __future__ import annotations

import io
import re
import tempfile
import unicodedata
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from django.core.files.base import ContentFile

from ..models import Destinatario, LoteImportacao, RelatorioComissionamento
from ..parceiros import normalizar_razao
from tickets.models import Parceiro


def _norm_texto(v) -> str:
    txt = str(v or "").strip().upper()
    txt = unicodedata.normalize("NFKD", txt)
    return "".join(c for c in txt if not unicodedata.combining(c))


def _sanitizar_nome(nome: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "", str(nome or "").strip()) or "PDV"


def extrair_razoes_sociais(texto: str) -> list[str]:
    partes = re.split(r"[;\n\r]+", str(texto or ""))
    return [p.strip() for p in partes if p.strip()]


def mapa_pdv_razoes() -> dict[int, dict]:
    """parceiro_id → {pdv_nome, razoes, email_especialista} a partir do cadastro do PDV (+ legado destinatários)."""
    mapa: dict[int, dict] = {}
    for parceiro in Parceiro.objects.filter(ativo=True).exclude(razao_social="").select_related("especialista"):
        email_esp = ""
        nome_esp = ""
        if parceiro.especialista:
            if parceiro.especialista.email:
                email_esp = parceiro.especialista.email.strip()
            nome_esp = (parceiro.especialista.get_full_name() or parceiro.especialista.username).strip()
        mapa[parceiro.id] = {
            "pdv_nome": parceiro.nome,
            "razoes": [parceiro.razao_social.strip()],
            "email_especialista": email_esp,
            "especialista_nome": nome_esp,
        }
    qs = (
        Destinatario.objects.filter(ativo=True, envio_comissionamento=True)
        .exclude(razoes_sociais_comissionamento="")
        .select_related("parceiro__especialista")
    )
    for dest in qs:
        extras = extrair_razoes_sociais(dest.razoes_sociais_comissionamento)
        if not extras:
            continue
        email_esp = ""
        nome_esp = ""
        if dest.parceiro and dest.parceiro.especialista:
            if dest.parceiro.especialista.email:
                email_esp = dest.parceiro.especialista.email.strip()
            nome_esp = (dest.parceiro.especialista.get_full_name() or dest.parceiro.especialista.username).strip()
        info = mapa.setdefault(
            dest.parceiro_id,
            {
                "pdv_nome": dest.parceiro.nome,
                "razoes": [],
                "email_especialista": email_esp,
                "especialista_nome": nome_esp,
            },
        )
        if not info.get("email_especialista") and email_esp:
            info["email_especialista"] = email_esp
        if not info.get("especialista_nome") and nome_esp:
            info["especialista_nome"] = nome_esp
        for razao in extras:
            if razao not in info["razoes"]:
                info["razoes"].append(razao)
    for info in mapa.values():
        info["razoes"] = sorted(info["razoes"])
    return {pid: info for pid, info in mapa.items() if info["razoes"]}


def _resolver_aba(nome_alvo: str, abas) -> str | None:
    alvo = _norm_texto(nome_alvo)
    for aba in abas:
        if _norm_texto(aba) == alvo:
            return aba
    return None


def _resolver_coluna_razao(df: pd.DataFrame):
    for col in df.columns:
        col_norm = _norm_texto(col).replace("_", " ")
        if col_norm == "RAZAO SOCIAL":
            return col
    return None


def _carregar_aba(xls: pd.ExcelFile, sheet_name: str, candidatos: list[str], max_linhas: int = 20):
    df_bruto = pd.read_excel(xls, sheet_name=sheet_name, header=None)
    alvos = {_norm_texto(c).replace("_", " ") for c in candidatos}
    linha_header = None
    for i in range(min(max_linhas, len(df_bruto))):
        valores_norm = {_norm_texto(v).replace("_", " ") for v in df_bruto.iloc[i].tolist()}
        if any(a in valores_norm for a in alvos):
            linha_header = i
            break
    if linha_header is None:
        return pd.read_excel(xls, sheet_name=sheet_name)
    df = pd.read_excel(xls, sheet_name=sheet_name, header=linha_header)
    df = df.loc[:, [str(c).strip() != "" and not str(c).startswith("Unnamed") for c in df.columns]]
    return df


def _find_col(df: pd.DataFrame, aliases: list[str]):
    mapa = {_norm_texto(c).replace("_", " "): c for c in df.columns}
    for a in aliases:
        col = mapa.get(_norm_texto(a).replace("_", " "))
        if col:
            return col
    return None


def _to_float(valor) -> float:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    txt = str(valor).strip()
    if not txt:
        return 0.0
    txt = txt.replace("R$", "").replace(".", "").replace(",", ".")
    try:
        return float(txt)
    except Exception:
        return 0.0


def _fmt_moeda(v: float) -> str:
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _classificar_coluna(nome_coluna: str) -> str:
    norm = _norm_texto(str(nome_coluna)).replace("_", " ")
    if any(k in norm for k in ["VALOR", "COMISSAO", "TOTAL", "PRECO", "RECEITA", "FATURAMENTO", "VALOR UNITARIO"]):
        return "moeda"
    if any(k in norm for k in ["DATA", "DT ", "EMISSAO", "ATIVACAO", "TRANSICAO", "CANCELAMENTO", "CRIADO"]):
        return "data"
    if any(k in norm for k in ["CNPJ", "CPF", "HANA", "FORNECEDOR", "DOCUMENTO", "ITEM", "CENTRO", "CONTRATO", "MATRICULA", "ORDEM", "PEDIDO", "CEP", "TELEFONE"]):
        return "codigo"
    return "texto"


def gerar_planilha_comissionamento_formatada(pedido_df: pd.DataFrame, linha_df: pd.DataFrame) -> bytes:
    """Gera workbook XLSX com apresentação executiva: cabeçalhos navy, bordas, zebra striping e formatação de células."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws_pedido = wb.active
    ws_pedido.title = "PEDIDO"
    ws_linha = wb.create_sheet(title="LINHA_A_LINHA")

    fill_header = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    font_data = Font(name="Calibri", size=10, color="1E293B")

    borda_fina = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    def _formatar_aba(ws, df: pd.DataFrame):
        ws.views.sheetView[0].showGridLines = True
        ws.freeze_panes = "A2"

        if df is None or df.empty:
            ws.append(["Nenhum registro para este PDV."])
            return

        cols = list(df.columns)
        ws.append(cols)
        ws.row_dimensions[1].height = 28
        tipos_col = [_classificar_coluna(c) for c in cols]

        for col_idx in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = borda_fina

        max_larguras = {i: len(str(c)) for i, c in enumerate(cols, start=1)}

        for row_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
            ws.row_dimensions[row_idx].height = 20
            is_zebra = (row_idx % 2 == 1)

            for col_idx, (_, val) in enumerate(zip(cols, row_data), start=1):
                tipo = tipos_col[col_idx - 1]
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_data
                cell.border = borda_fina
                if is_zebra:
                    cell.fill = fill_zebra

                if pd.isna(val) or val is None:
                    cell.value = ""
                    continue

                tam_str = 10
                if tipo == "moeda":
                    num_val = _to_float(val)
                    cell.value = num_val
                    cell.number_format = "R$ #,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    tam_str = len(f"R$ {num_val:,.2f}")
                elif tipo == "data":
                    if isinstance(val, (datetime, pd.Timestamp)):
                        cell.value = val.to_pydatetime().replace(tzinfo=None) if hasattr(val, "to_pydatetime") else val
                        cell.number_format = "DD/MM/YYYY"
                    elif isinstance(val, date):
                        cell.value = val
                        cell.number_format = "DD/MM/YYYY"
                    else:
                        cell.value = str(val).strip()
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    tam_str = 12
                elif tipo == "codigo":
                    val_str = str(val).strip()
                    if val_str.endswith(".0"):
                        val_str = val_str[:-2]
                    cell.value = val_str
                    cell.number_format = "@"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    tam_str = len(val_str)
                else:
                    if isinstance(val, (int, float)) and not isinstance(val, bool):
                        cell.value = val
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        tam_str = len(str(val))
                    else:
                        val_str = str(val).strip()
                        cell.value = val_str
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        tam_str = len(val_str)

                if tam_str > max_larguras.get(col_idx, 0):
                    max_larguras[col_idx] = tam_str

        for col_idx in range(1, len(cols) + 1):
            col_letter = get_column_letter(col_idx)
            largura = max(13, min(max_larguras.get(col_idx, 10) + 4, 50))
            ws.column_dimensions[col_letter].width = largura

    _formatar_aba(ws_pedido, pedido_df)
    _formatar_aba(ws_linha, linha_df)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def montar_mensagem(
    caminho_anexo: Path | str,
    pdv_nome: str,
    limite_pedidos: int = 15,
    email_copia: str = "",
) -> tuple[str, float, float]:
    path = Path(caminho_anexo)
    engine = "pyxlsb" if path.suffix.lower() == ".xlsb" else None
    try:
        pedido_df = pd.read_excel(path, sheet_name="PEDIDO", engine=engine)
        linha_df = pd.read_excel(path, sheet_name="LINHA_A_LINHA", engine=engine)
    except Exception as exc:
        return (
            f"📁 *Comissionamento por PDV*\nPDV: *{pdv_nome}*\n\n"
            f"Não foi possível montar resumo detalhado: {exc}",
            0.0,
            0.0,
        )

    c_pedido = _find_col(pedido_df, ["DOCUMENTO DE COMPRAS", "DOCUMENTO"])
    c_item = _find_col(pedido_df, ["ITEM"])
    c_valor = _find_col(pedido_df, ["VALOR"])
    c_hana = _find_col(pedido_df, ["FORNECEDOR"])
    c_razao = _find_col(pedido_df, ["RAZÃO SOCIAL", "RAZAO SOCIAL", "RAZAO_SOCIAL"])
    c_cnpj = _find_col(pedido_df, ["CNPJ"])
    c_canal = _find_col(pedido_df, ["CANAL"])
    c_centro = _find_col(pedido_df, ["CENTRO"])
    c_ciclo = _find_col(pedido_df, ["CICLO"])
    c_sub_evento = _find_col(linha_df, ["SUB_EVENTO", "SUB EVENTO"])
    c_comissao = _find_col(linha_df, ["COMISSAO", "COMISSÃO"])

    msg = [f"📁 *Comissionamento por PDV*", f"PDV: *{pdv_nome}*"]
    total_valor_pedido = 0.0

    if c_pedido and c_item and c_valor:
        limite = min(len(pedido_df), max(1, int(limite_pedidos)))
        for i in range(limite):
            row = pedido_df.iloc[i]
            valor = _to_float(row.get(c_valor))
            total_valor_pedido += valor
            msg.extend(
                [
                    "",
                    f"Pedido: {row.get(c_pedido, '-')}",
                    f"ITEM: {row.get(c_item, '-')}",
                    f"VALOR: {_fmt_moeda(valor)}",
                    f"CÓDIGO HANA: {row.get(c_hana, '-') if c_hana else '-'}",
                    f"RAZÃO SOCIAL: {row.get(c_razao, '-') if c_razao else '-'}",
                    f"CNPJ: {row.get(c_cnpj, '-') if c_cnpj else '-'}",
                    f"CANAL: {row.get(c_canal, '-') if c_canal else '-'}",
                    f"CENTRO: {row.get(c_centro, '-') if c_centro else '-'}",
                    f"REFERÊNCIA: COMISSAO {row.get(c_ciclo, '-') if c_ciclo else '-'}",
                ]
            )
        if len(pedido_df) > limite:
            msg.append(f"\n... e mais {len(pedido_df) - limite} pedido(s) no anexo.")
    else:
        msg.append("\nNão foi possível montar bloco PEDIDO completo (colunas ausentes).")

    total_comissao = 0.0
    if c_sub_evento and c_comissao and not linha_df.empty:
        linha_tmp = linha_df[[c_sub_evento, c_comissao]].copy()
        linha_tmp[c_comissao] = linha_tmp[c_comissao].apply(_to_float)
        resumo = linha_tmp.groupby(c_sub_evento, dropna=False)[c_comissao].sum().sort_values(ascending=False)
        total_comissao = float(resumo.sum())
        msg.append("\nResumo LINHA_A_LINHA por SUB_EVENTO:")
        for sub_evento, soma in resumo.items():
            msg.append(f"- {sub_evento}: {_fmt_moeda(float(soma))}")
        msg.append(f"\nTOTAL LINHA_A_LINHA: {_fmt_moeda(total_comissao)}")
        msg.append(f"TOTAL PEDIDO: {_fmt_moeda(total_valor_pedido)}")
        diferenca = total_valor_pedido - total_comissao
        status = "OK" if abs(diferenca) < 0.01 else f"Diferença de {_fmt_moeda(diferenca)}"
        msg.append(f"Conferência total: {status}")
    else:
        msg.append("\nNão foi possível montar resumo de SUB_EVENTO/COMISSAO (colunas ausentes).")

    empresa_assunto = ""
    ciclo_assunto = ""
    if not pedido_df.empty:
        if c_razao:
            s_razao = pedido_df[c_razao].dropna()
            if not s_razao.empty:
                empresa_assunto = str(s_razao.iloc[0]).strip()
        if c_ciclo:
            s_ciclo = pedido_df[c_ciclo].dropna()
            if not s_ciclo.empty:
                raw_c = str(s_ciclo.iloc[0]).strip()
                ciclo_assunto = re.sub(r"^COMISS[ÃA]O\s*", "", raw_c, flags=re.IGNORECASE).strip()

    empresa_final = empresa_assunto or pdv_nome
    assunto_email = f"{empresa_final}_{ciclo_assunto}" if ciclo_assunto else f"{empresa_final}_[CICLO]"

    email_dest = (email_copia or "").strip() or "rogerio.pacheco@niointernet.com.br"
    msg.append(
        "\nOrientação para envio do email: \n\n"
        "Enviar o email para recebimentonfes@niointernet.com.br\n"
        f"Com cópia para: {email_dest} e PP-GestaodosParceiros@niointernet.com.br\n\n"
        "No corpo do email retirar assinatura e não escrever nada no corpo do email \n"
        f"E o assunto do email deverá ser {assunto_email}"
    )

    return "\n".join(msg), total_valor_pedido, total_comissao


def _salvar_bytes_temporario(arquivo, nome: str) -> Path:
    conteudo = arquivo.read() if hasattr(arquivo, "read") else Path(arquivo).read_bytes()
    if hasattr(arquivo, "seek"):
        arquivo.seek(0)
    sufixo = Path(nome).suffix.lower() or ".xlsx"
    tmp = tempfile.NamedTemporaryFile(suffix=sufixo, delete=False)
    tmp.write(conteudo)
    tmp.close()
    return Path(tmp.name)


def processar_comissionamento(arquivo, nome: str, lote: LoteImportacao) -> dict:
    """Separa ciclo (PEDIDO + LINHA_A_LINHA) por PDV conforme razões dos destinatários."""
    mapa = mapa_pdv_razoes()
    if not mapa:
        raise ValueError(
            "Nenhum PDV com razão social cadastrada em Parceiros "
            "(ou em Destinatários, legado)."
        )

    caminho = _salvar_bytes_temporario(arquivo, nome)
    xls = None
    try:
        engine = "pyxlsb" if caminho.suffix.lower() == ".xlsb" else None
        try:
            xls = pd.ExcelFile(caminho, engine=engine)
        except Exception as exc:
            raise ValueError(f"Não foi possível abrir o arquivo: {exc}") from exc

        aba_pedido = _resolver_aba("PEDIDO", xls.sheet_names)
        aba_linha = _resolver_aba("LINHA_A_LINHA", xls.sheet_names)
        if not aba_pedido or not aba_linha:
            raise ValueError(
                f"Abas PEDIDO e LINHA_A_LINHA obrigatórias. Encontradas: {xls.sheet_names}"
            )

        candidatos = ["RAZÃO SOCIAL", "RAZAO SOCIAL", "RAZAO_SOCIAL"]
        df_pedido = _carregar_aba(xls, aba_pedido, candidatos)
        df_linha = _carregar_aba(xls, aba_linha, candidatos)
        col_razao_pedido = _resolver_coluna_razao(df_pedido)
        col_razao_linha = _resolver_coluna_razao(df_linha)
        if not col_razao_pedido or not col_razao_linha:
            raise ValueError("Coluna 'RAZÃO SOCIAL' não encontrada em PEDIDO ou LINHA_A_LINHA.")

        # Fecha handle do arquivo antes de gerar saídas / limpeza (Windows)
        xls.close()
        xls = None

        base_nome = Path(nome).stem
        gerados = 0
        sem_linhas = 0

        for parceiro_id, info in mapa.items():
            pdv_nome = info["pdv_nome"]
            razoes_norm = {normalizar_razao(r) for r in info["razoes"]}
            mask_pedido = df_pedido[col_razao_pedido].apply(lambda v: normalizar_razao(str(v))).isin(
                razoes_norm
            )
            mask_linha = df_linha[col_razao_linha].apply(lambda v: normalizar_razao(str(v))).isin(
                razoes_norm
            )
            pedido_pdv = df_pedido[mask_pedido].copy()
            linha_pdv = df_linha[mask_linha].copy()
            if pedido_pdv.empty and linha_pdv.empty:
                sem_linhas += 1
                continue

            conteudo_xlsx = gerar_planilha_comissionamento_formatada(pedido_pdv, linha_pdv)
            nome_saida = f"{base_nome}_{_sanitizar_nome(pdv_nome)}.xlsx"

            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_out:
                tmp_out.write(conteudo_xlsx)
                caminho_out = Path(tmp_out.name)
            try:
                mensagem, total_pedido, total_comissao = montar_mensagem(
                    caminho_out, pdv_nome, email_copia=info.get("email_especialista", "")
                )
            finally:
                caminho_out.unlink(missing_ok=True)

            rel = RelatorioComissionamento(
                lote=lote,
                parceiro_id=parceiro_id,
                pdv_nome=pdv_nome,
                qtd_pedido=int(len(pedido_pdv)),
                qtd_linha=int(len(linha_pdv)),
                total_pedido=total_pedido,
                total_comissao=total_comissao,
                mensagem=mensagem,
                detalhes={"razoes": info["razoes"]},
            )
            rel.arquivo.save(nome_saida, ContentFile(conteudo_xlsx), save=True)
            gerados += 1

        return {
            "pdvs": gerados,
            "sem_linhas": sem_linhas,
            "pdvs_configurados": len(mapa),
            "arquivo": nome,
        }
    finally:
        if xls is not None:
            try:
                xls.close()
            except Exception:
                pass
        try:
            caminho.unlink(missing_ok=True)
        except PermissionError:
            pass
