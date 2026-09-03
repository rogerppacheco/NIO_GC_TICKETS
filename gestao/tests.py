from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from unittest import skipUnless

from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook

from tickets.models import Parceiro, PerfilStaff

from gestao.models import (
    CadastroTerceiro,
    GrossMensal,
    HistoricoChurn,
    HistoricoOSAB,
    RelatorioFPD,
    VendaOSAB,
)
from gestao.parceiros import resolver_parceiro_id
from gestao.periodo import periodo_ativo, salvar_periodo
from gestao.pipelines.churn import processar_churn
from gestao.pipelines.fpd import processar_fpd
from gestao.pipelines.osab import processar_osab
from gestao.terceiros import importar_sysmap
from gestao.models import LoteImportacao


def _xlsx(linhas, colunas):
    wb = Workbook()
    ws = wb.active
    ws.append(colunas)
    for linha in linhas:
        ws.append(linha)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "base.xlsx"
    return buf


def _gdp_xlsx(linhas, aba="PAP (Local)", nome="gdp.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = aba
    ws.append(["UF", "MUNICIPIO", "PORTFOLIO_GDP_20_08", "COD_IBGE"])
    for linha in linhas:
        ws.append(linha)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = nome
    return buf


class ParceiroMatchTests(TestCase):
    def setUp(self):
        self.pdv = Parceiro.objects.create(codigo_pdv="1068281", nome="INOVA MG")

    def test_alias_razao_social(self):
        self.assertEqual(
            resolver_parceiro_id("LUISA SERVICOS DE TELEFONIA MOVEL LTDA"),
            self.pdv.id,
        )

    def test_nome_exato(self):
        self.assertEqual(resolver_parceiro_id("INOVA MG"), self.pdv.id)

    def test_razao_social_cadastrada_no_parceiro(self):
        pdv = Parceiro.objects.create(
            codigo_pdv="kk",
            nome="K&K",
            razao_social="KEK TELECOM INTERMEDIACAO LTDA",
        )
        self.assertEqual(
            resolver_parceiro_id("KEK TELECOM INTERMEDIACAO LTDA"),
            pdv.id,
        )


class SysmapImportTests(TestCase):
    def setUp(self):
        self.pdv = Parceiro.objects.create(codigo_pdv="1", nome="INOVA MG")

    def test_upsert_por_chave(self):
        arquivo = _xlsx(
            [
                [
                    "LUISA SERVICOS DE TELEFONIA MOVEL",
                    "ANA",
                    "123",
                    "a@x.com",
                    "TT1",
                    "CLT",
                    "VENDEDOR",
                    "Ativo",
                    "Ativo",
                    "Alocado",
                    "01/01/2025",
                    "",
                    "",
                ]
            ],
            [
                "Razão Social",
                "Terceiro",
                "CPF",
                "Email",
                "Chave de Acesso",
                "Vínculo",
                "Cargo/Função",
                "Situação Terceiro Empresa",
                "Situação Funcional",
                "Situação Terceiro Contrato",
                "Data Alocação",
                "Data Desalocação",
                "Data Inativação",
            ],
        )
        resumo = importar_sysmap(arquivo, "terceiros.xlsx")
        self.assertEqual(resumo["inseridos"], 1)
        t = CadastroTerceiro.objects.get(chave_acesso="TT1")
        self.assertTrue(t.ativo)
        self.assertEqual(t.parceiro_id, self.pdv.id)
        self.assertEqual(t.cargo_funcao, "VENDEDOR")

    def test_desativado_nao_elegivel_capilaridade(self):
        from gestao.terceiros import terceiro_elegivel_capilaridade

        self.assertFalse(
            terceiro_elegivel_capilaridade("Desativado", "Desativado", "Desalocado")
        )
        self.assertFalse(
            terceiro_elegivel_capilaridade("Desativado", "Ativo", "Alocado")
        )

    def test_consolidacao_prefere_linha_desativada_recente(self):
        arquivo = _xlsx(
            [
                [
                    "LUISA SERVICOS DE TELEFONIA MOVEL",
                    "JOAO",
                    "123",
                    "a@x.com",
                    "TT774192",
                    "CLT",
                    "VENDEDOR",
                    "Ativo",
                    "Ativo",
                    "Alocado",
                    "01/01/2024",
                    "",
                    "",
                ],
                [
                    "LUISA SERVICOS DE TELEFONIA MOVEL",
                    "JOAO",
                    "123",
                    "a@x.com",
                    "TT774192",
                    "CLT",
                    "VENDEDOR",
                    "Desativado",
                    "Desativado",
                    "Desalocado",
                    "01/01/2024",
                    "15/08/2025",
                    "15/08/2025",
                ],
            ],
            [
                "Razão Social",
                "Terceiro",
                "CPF",
                "Email",
                "Chave de Acesso",
                "Vínculo",
                "Cargo/Função",
                "Situação Terceiro Empresa",
                "Situação Funcional",
                "Situação Terceiro Contrato",
                "Data Alocação",
                "Data Desalocação",
                "Data Inativação",
            ],
        )
        importar_sysmap(arquivo, "terceiros.xlsx")
        t = CadastroTerceiro.objects.get(chave_acesso="TT774192")
        self.assertEqual(t.situacao_empresa, "Desativado")
        self.assertFalse(t.ativo)

    def test_desativado_nao_entra_na_capilaridade(self):
        from gestao.pipelines.osab import linhas_capilaridade_pdv

        pdv = Parceiro.objects.create(codigo_pdv="2", nome="RECORD")
        CadastroTerceiro.objects.create(
            chave_acesso="TT816374",
            nome_terceiro="MARIA",
            razao_social="RECORD",
            parceiro=pdv,
            cargo_funcao="VENDEDOR",
            situacao_empresa="Desativado",
            situacao_funcional="Desativado",
            situacao_contrato="Desalocado",
            ativo=False,
        )
        chaves = {l["matricula_vendedor"] for l in linhas_capilaridade_pdv(pdv)}
        self.assertNotIn("TT816374", chaves)


class OsabCapilaridadeTests(TestCase):
    def setUp(self):
        self.pdv = Parceiro.objects.create(codigo_pdv="1", nome="RECORD")
        CadastroTerceiro.objects.create(
            chave_acesso="TT99",
            nome_terceiro="JOAO",
            razao_social="RECORD",
            parceiro=self.pdv,
            cargo_funcao="VENDEDOR",
            situacao_empresa="Ativo",
            situacao_funcional="Ativo",
            situacao_contrato="Alocado",
            ativo=True,
        )

    def test_importa_venda_e_gera_capilaridade(self):
        hoje = timezone.localdate()
        abertura = datetime(hoje.year, hoje.month, 1, 10, 0)
        arquivo = _xlsx(
            [
                [
                    "P1",
                    abertura,
                    "TT99",
                    "JOAO",
                    "RECORD",
                    abertura,
                    abertura,
                    "Concluído",
                    "500 MEGA",
                ]
            ],
            [
                "PEDIDO",
                "DT_REF",
                "MATRICULA_VENDEDOR",
                "NOME_VENDEDOR",
                "DESCRICAO",
                "DATA_ABERTURA",
                "DATA_FECHAMENTO",
                "SITUACAO",
                "VELOCIDADE",
            ],
        )
        resumo = processar_osab(arquivo, "OSAB.xlsx", hoje.year, hoje.month)
        self.assertEqual(resumo["vendas"]["inseridos"], 1)
        self.assertEqual(VendaOSAB.objects.get(pedido="P1").parceiro_id, self.pdv.id)
        self.assertGreaterEqual(resumo["capilaridade"]["linhas"], 1)
        self.assertTrue(HistoricoOSAB.objects.filter(parceiro=self.pdv).exists())

    def test_osab_formato_bi(self):
        hoje = timezone.localdate()
        abertura = datetime(hoje.year, hoje.month, 1, 10, 0)
        arquivo = _xlsx(
            [
                [
                    "P-BI",
                    abertura,
                    "TT99",
                    "RECORD",
                    abertura,
                    abertura,
                    "Concluído",
                    "500MB",
                    "MANAUS",
                    "BL_500MB_VAR",
                    "PP",
                    "CARLA JULIANE",
                    1069075,
                    "Boleto Digital",
                    "APPPAP",
                ],
                [
                    "Filtros aplicados:\nANOMES",
                    abertura,
                    "",
                    "",
                    abertura,
                    abertura,
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "Filtros aplicados:\nANOMES",
                ],
            ],
            [
                "PEDIDO",
                "DT_REF",
                "MATRICULA_VENDEDOR",
                "nm_pdv_rel",
                "DATA_ABERTURA",
                "DATA_FECHAMENTO",
                "SITUACAO",
                "VELOCIDADE",
                "LOCALIDADE",
                "CAMPANHA",
                "GERENCIA",
                "nm_gc",
                "PDV_SAP",
                "meio_pagamento",
                "PRODUTO",
            ],
        )
        resumo = processar_osab(arquivo, "OSAB-BI.xlsx", hoje.year, hoje.month)
        self.assertEqual(resumo["vendas"]["inseridos"], 1)
        venda = VendaOSAB.objects.get(pedido="P-BI")
        self.assertEqual(venda.pdv_nome, "RECORD")
        self.assertEqual(venda.municipio, "MANAUS")
        self.assertEqual(venda.oferta, "BL_500MB_VAR")
        self.assertEqual(venda.nm_gc, "CARLA JULIANE")
        self.assertEqual(venda.gerencia, "PP")
        self.assertEqual(venda.pdv_sap, "1069075")
        self.assertEqual(venda.meio_pagamento, "Boleto Digital")
        self.assertEqual(venda.nome_vendedor, "TT99")

    def test_vendedor_externo_entra_na_mascara(self):
        from gestao.pipelines.osab import linhas_capilaridade_pdv

        CadastroTerceiro.objects.create(
            chave_acesso="TT100",
            nome_terceiro="MARIA EXTERNA",
            razao_social="RECORD",
            parceiro=self.pdv,
            cargo_funcao="VENDEDOR EXTERNO",
            situacao_empresa="Ativo",
            situacao_funcional="Ativo",
            situacao_contrato="Alocado",
            ativo=True,
        )
        linhas = linhas_capilaridade_pdv(self.pdv)
        chaves = {l["matricula_vendedor"] for l in linhas}
        self.assertIn("TT99", chaves)
        self.assertIn("TT100", chaves)

    def test_vendedor_vinculado_por_osab_sem_razao_social(self):
        from gestao.pipelines.osab import linhas_capilaridade_pdv

        pdv_kk = Parceiro.objects.create(codigo_pdv="kk", nome="K&K")
        CadastroTerceiro.objects.create(
            chave_acesso="TT815070",
            nome_terceiro="VENDEDOR KK",
            razao_social="KEK TELECOM INTERMEDIACAO LTDA",
            parceiro=None,
            cargo_funcao="VENDEDOR",
            situacao_empresa="Ativo",
            situacao_funcional="Ativo",
            situacao_contrato="Alocado",
            ativo=True,
        )
        VendaOSAB.objects.create(
            pedido="P-KK-1",
            pdv_nome="K&K",
            matricula_vendedor="TT815070",
            nome_vendedor="VENDEDOR KK",
            data_abertura=timezone.now(),
            situacao="Concluído",
        )
        chaves = {l["matricula_vendedor"] for l in linhas_capilaridade_pdv(pdv_kk)}
        self.assertIn("TT815070", chaves)

    def test_vendedor_sem_osab_continua_por_razao_social(self):
        from gestao.pipelines.osab import linhas_capilaridade_pdv

        pdv = Parceiro.objects.create(codigo_pdv="3", nome="INOVA MG")
        CadastroTerceiro.objects.create(
            chave_acesso="TT300",
            nome_terceiro="ANA NOVA",
            razao_social="LUISA SERVICOS DE TELEFONIA MOVEL LTDA",
            parceiro=None,
            cargo_funcao="VENDEDOR",
            situacao_empresa="Ativo",
            situacao_funcional="Ativo",
            situacao_contrato="Alocado",
            ativo=True,
        )
        chaves = {l["matricula_vendedor"] for l in linhas_capilaridade_pdv(pdv)}
        self.assertIn("TT300", chaves)

    def test_desativado_nao_entra_mesmo_com_osab(self):
        from gestao.pipelines.osab import linhas_capilaridade_pdv

        pdv_kk = Parceiro.objects.create(codigo_pdv="kk2", nome="K&K 2")
        CadastroTerceiro.objects.create(
            chave_acesso="TT816374",
            nome_terceiro="MARIA OFF",
            razao_social="KEK TELECOM INTERMEDIACAO LTDA",
            parceiro=None,
            cargo_funcao="VENDEDOR",
            situacao_empresa="Desativado",
            situacao_funcional="Desativado",
            situacao_contrato="Desalocado",
            ativo=False,
        )
        VendaOSAB.objects.create(
            pedido="P-KK-2",
            pdv_nome="K&K 2",
            matricula_vendedor="TT816374",
            data_abertura=timezone.now(),
        )
        chaves = {l["matricula_vendedor"] for l in linhas_capilaridade_pdv(pdv_kk)}
        self.assertNotIn("TT816374", chaves)

    def test_mascara_desativada_com_venda_no_mes(self):
        from gestao.relatorios import montar_mascara_pdv

        CadastroTerceiro.objects.create(
            chave_acesso="TT807399",
            nome_terceiro="ANDRE JUNIO DE OLIVEIRA SANTOS SOUZA",
            razao_social="RECORD",
            parceiro=self.pdv,
            cargo_funcao="VENDEDOR",
            situacao_empresa="Desativado",
            situacao_funcional="Desativado",
            situacao_contrato="Desalocado",
            ativo=False,
        )
        agora = timezone.now()
        VendaOSAB.objects.create(
            pedido="P-DESAT",
            pdv_nome="RECORD",
            matricula_vendedor="TT807399",
            data_abertura=agora,
            parceiro=self.pdv,
        )
        mascara = montar_mascara_pdv(self.pdv, agora.year, agora.month)
        self.assertIn("TTs desativadas com vendas no mês", mascara)
        self.assertIn("TT807399 · ANDRE SOUZA", mascara)
        recuperar = mascara.split("📴")[0]
        self.assertNotIn("TT807399", recuperar)

    def test_filtro_tt_nome_cargo(self):
        from gestao.pipelines.osab import linhas_capilaridade_pdv

        CadastroTerceiro.objects.create(
            chave_acesso="TT200",
            nome_terceiro="CARLOS SUPER",
            razao_social="RECORD",
            parceiro=self.pdv,
            cargo_funcao="SUPERVISOR",
            situacao_empresa="Ativo",
            situacao_funcional="Ativo",
            situacao_contrato="Alocado",
            ativo=True,
        )
        so_tt = linhas_capilaridade_pdv(self.pdv, {"tt": "TT99"})
        self.assertEqual([l["matricula_vendedor"] for l in so_tt], ["TT99"])
        so_nome = linhas_capilaridade_pdv(self.pdv, {"nome": "JOAO"})
        self.assertEqual([l["matricula_vendedor"] for l in so_nome], ["TT99"])
        so_cargo = linhas_capilaridade_pdv(self.pdv, {"cargo": "SUPERVISOR"})
        self.assertEqual([l["matricula_vendedor"] for l in so_cargo], ["TT200"])
        padrao = linhas_capilaridade_pdv(self.pdv)
        self.assertEqual([l["matricula_vendedor"] for l in padrao], ["TT99"])

    def test_mascara_usa_primeiro_e_ultimo_nome(self):
        from gestao.relatorios import _linha_tt, primeiro_ultimo_nome

        self.assertEqual(
            primeiro_ultimo_nome("CAUAN HENRIQUE DE OLIVEIRA DA CRUZ"), "CAUAN CRUZ"
        )
        self.assertEqual(primeiro_ultimo_nome("ANA PAULA"), "ANA PAULA")
        self.assertEqual(primeiro_ultimo_nome("JOAO"), "JOAO")
        linha = _linha_tt("TT99", 2, timezone.now(), "JOAO VITOR VIEIRA DIAS")
        self.assertIn("JOAO DIAS", linha)
        self.assertNotIn("VITOR", linha)

    def test_mascara_auditoria_usa_nome_sysmap_e_titulo_recuperar(self):
        from gestao.relatorios import montar_mascara_pdv
        from gestao.terceiros import CARGO_OPERADOR_VENDA_INTERNA

        CadastroTerceiro.objects.create(
            chave_acesso="TT822566",
            nome_terceiro="MARIA SILVA SANTOS",
            razao_social="RECORD",
            parceiro=self.pdv,
            cargo_funcao=CARGO_OPERADOR_VENDA_INTERNA,
            situacao_empresa="Ativo",
            situacao_funcional="Ativo",
            situacao_contrato="Alocado",
            ativo=True,
        )
        agora = timezone.now()
        VendaOSAB.objects.create(
            pedido="OP1",
            pdv_nome="RECORD",
            matricula_vendedor="TT822566",
            data_abertura=agora,
            parceiro=self.pdv,
        )
        mascara = montar_mascara_pdv(self.pdv, agora.year, agora.month)
        self.assertIn("*Recuperar:", mascara)
        self.assertNotIn("Inativos a recuperar", mascara)
        self.assertIn("Vendas Operador de Vendas interna", mascara)
        self.assertIn("TT822566 · MARIA SANTOS", mascara)

    def test_planilha_capilaridade_aceita_datetime_com_fuso(self):
        from gestao.planilhas import planilha_capilaridade

        VendaOSAB.objects.create(
            pedido="P-TZ",
            pdv_nome="RECORD",
            matricula_vendedor="TT99",
            data_abertura=timezone.now(),
            parceiro=self.pdv,
        )
        dados, nome = planilha_capilaridade(self.pdv)
        self.assertTrue(dados.startswith(b"PK"))
        self.assertIn("RECORD", nome)

    def test_planilha_capilaridade_com_nat_e_venda(self):
        import pandas as pd

        from gestao.planilhas import _celula_excel, df_para_xlsx

        self.assertIsNone(_celula_excel(pd.NaT))
        CadastroTerceiro.objects.create(
            chave_acesso="TT101",
            nome_terceiro="SEM VENDA",
            razao_social="RECORD",
            parceiro=self.pdv,
            cargo_funcao="VENDEDOR",
            situacao_empresa="Ativo",
            situacao_funcional="Ativo",
            situacao_contrato="Alocado",
            ativo=True,
        )
        VendaOSAB.objects.create(
            pedido="P-MIX",
            pdv_nome="RECORD",
            matricula_vendedor="TT99",
            data_abertura=timezone.now(),
            parceiro=self.pdv,
        )
        from gestao.planilhas import planilha_capilaridade

        dados, _ = planilha_capilaridade(self.pdv)
        self.assertTrue(dados.startswith(b"PK"))
        df = pd.DataFrame({"Última venda": [timezone.now(), pd.NaT, None]})
        dados2, _ = df_para_xlsx(df, "mix.xlsx")
        self.assertTrue(dados2.startswith(b"PK"))


class CadastroParceirosOsabTests(TestCase):
    def setUp(self):
        self.existente = Parceiro.objects.create(codigo_pdv="1068281", nome="INOVA MG")
        self.orfao = Parceiro.objects.create(codigo_pdv="999", nome="PDV ANTIGO")

    def test_cria_faltantes_e_nao_exclui(self):
        from gestao.parceiros import sincronizar_parceiros_osab

        agora = timezone.now()
        VendaOSAB.objects.create(pedido="A", pdv_nome="INOVA MG", data_abertura=agora)
        VendaOSAB.objects.create(pedido="B", pdv_nome="NOVO PDV", data_abertura=agora)
        cad = sincronizar_parceiros_osab()
        self.assertIn("INOVA MG", cad["ja_ok"])
        self.assertEqual(cad["criados"], ["NOVO PDV"])
        self.assertIn("PDV ANTIGO", cad["nio_sem_osab"])
        self.assertTrue(Parceiro.objects.filter(pk=self.existente.pk).exists())
        self.assertTrue(
            Parceiro.objects.filter(pk=self.orfao.pk, nome="PDV ANTIGO").exists()
        )
        novo = Parceiro.objects.get(nome="NOVO PDV")
        self.assertTrue(novo.codigo_pdv.startswith("OSAB-"))
        self.assertEqual(VendaOSAB.objects.get(pedido="B").parceiro_id, novo.id)

    def test_define_especialista_pelo_nm_gc(self):
        from gestao.parceiros import formatar_nome_pessoa, sincronizar_parceiros_osab

        self.assertEqual(formatar_nome_pessoa("JOAO"), "Joao")
        self.assertEqual(formatar_nome_pessoa("ANA PAULA"), "Ana Paula")
        agora = timezone.now()
        VendaOSAB.objects.create(
            pedido="C",
            pdv_nome="PDV GC",
            nm_gc="MARIA FERNANDA",
            gerencia="MG INTERIOR",
            data_abertura=agora,
        )
        cad = sincronizar_parceiros_osab()
        pdv = Parceiro.objects.get(nome="PDV GC")
        self.assertEqual(pdv.especialista.first_name, "Maria Fernanda")
        self.assertEqual(pdv.especialista.perfil_staff.gerencia, "MG INTERIOR")
        self.assertIn("Maria Fernanda", cad["especialistas_novos"])
        cad2 = sincronizar_parceiros_osab()
        self.assertEqual(cad2["criados"], [])
        VendaOSAB.objects.create(
            pedido="D", pdv_nome="PDV GC 2", nm_gc="maria fernanda", data_abertura=agora
        )
        sincronizar_parceiros_osab()
        pdv2 = Parceiro.objects.get(nome="PDV GC 2")
        self.assertEqual(pdv2.especialista_id, pdv.especialista_id)

    def test_associa_pdvs_do_gc_ao_especialista_homonimo(self):
        from gestao.parceiros import associar_parceiros_ao_especialista

        User = get_user_model()
        auto = User.objects.create_user(
            "marcella.oliveira.duarte",
            "auto@x.com",
            "x",
            is_staff=True,
            first_name="Marcella Oliveira Duarte",
        )
        PerfilStaff.objects.create(user=auto, papel=PerfilStaff.Papel.ESPECIALISTA)
        manual = User.objects.create_user(
            "Marcella",
            "ma@x.com",
            "x",
            is_staff=True,
            first_name="Marcella Oliveira Duarte",
        )
        PerfilStaff.objects.create(user=manual, papel=PerfilStaff.Papel.ESPECIALISTA)
        outro_esp = User.objects.create_user(
            "samuel", "sa@x.com", "x", is_staff=True, first_name="Samuel Octavio"
        )
        PerfilStaff.objects.create(user=outro_esp, papel=PerfilStaff.Papel.ESPECIALISTA)
        pdv = Parceiro.objects.create(
            codigo_pdv="DF1", nome="DIGITAL FIBRA", especialista=auto
        )
        outro = Parceiro.objects.create(
            codigo_pdv="RC1", nome="RECORD", especialista=outro_esp
        )
        agora = timezone.now()
        VendaOSAB.objects.create(
            pedido="M1",
            pdv_nome="DIGITAL FIBRA",
            nm_gc="MARCELLA OLIVEIRA DUARTE",
            data_abertura=agora,
        )
        VendaOSAB.objects.create(
            pedido="M2",
            pdv_nome="RECORD",
            nm_gc="SAMUEL OCTAVIO",
            data_abertura=agora,
        )
        movidos = associar_parceiros_ao_especialista(manual)
        pdv.refresh_from_db()
        outro.refresh_from_db()
        self.assertEqual(movidos, ["DIGITAL FIBRA"])
        self.assertEqual(pdv.especialista_id, manual.id)
        self.assertEqual(outro.especialista_id, outro_esp.id)

    def test_preenche_gerencia_de_especialista_existente(self):
        from gestao.parceiros import aplicar_gerencias_osab, sincronizar_parceiros_osab

        User = get_user_model()
        user = User.objects.create_user(
            "maria.fernanda", "mf@x.com", "x", is_staff=True, first_name="Maria Fernanda"
        )
        PerfilStaff.objects.create(user=user, papel=PerfilStaff.Papel.ESPECIALISTA)
        agora = timezone.now()
        VendaOSAB.objects.create(
            pedido="GER1",
            pdv_nome="PDV GER",
            nm_gc="MARIA FERNANDA",
            gerencia="MG INTERIOR",
            data_abertura=agora,
        )
        sincronizar_parceiros_osab()
        user.perfil_staff.refresh_from_db()
        self.assertEqual(user.perfil_staff.gerencia, "MG INTERIOR")
        pdv = Parceiro.objects.get(nome="PDV GER")
        self.assertEqual(pdv.especialista_id, user.id)
        aplicar_gerencias_osab()
        user.perfil_staff.refresh_from_db()
        self.assertEqual(user.perfil_staff.gerencia, "MG INTERIOR")

    def test_grafia_nao_cria_duplicata(self):
        from gestao.parceiros import sincronizar_parceiros_osab

        Parceiro.objects.create(codigo_pdv="1068432", nome="HF SERVIÇOS")
        cad = sincronizar_parceiros_osab(["HF SERVICOS"])
        self.assertEqual(cad["criados"], [])
        self.assertEqual(cad["grafia"][0]["cadastro"], "HF SERVIÇOS")
        self.assertEqual(Parceiro.objects.filter(nome__icontains="HF").count(), 1)

    def test_inativo_nao_e_recriado_nem_apagado(self):
        from gestao.parceiros import sincronizar_parceiros_osab

        self.existente.ativo = False
        self.existente.save(update_fields=["ativo"])
        cad = sincronizar_parceiros_osab(["INOVA MG"])
        self.assertEqual(cad["criados"], [])
        self.existente.refresh_from_db()
        self.assertFalse(self.existente.ativo)

    def test_cria_com_codigo_pdv_sap(self):
        from gestao.parceiros import sincronizar_parceiros_osab

        agora = timezone.now()
        VendaOSAB.objects.create(
            pedido="SAP1",
            pdv_nome="AMAZONTECH",
            pdv_sap="1071234",
            data_abertura=agora,
        )
        cad = sincronizar_parceiros_osab()
        self.assertEqual(cad["criados"], ["AMAZONTECH"])
        self.assertEqual(Parceiro.objects.get(nome="AMAZONTECH").codigo_pdv, "1071234")

    def test_placeholder_osab_vira_pdv_sap(self):
        from gestao.parceiros import sincronizar_parceiros_osab

        agora = timezone.now()
        Parceiro.objects.create(codigo_pdv="OSAB-AMAZONTECH", nome="AMAZONTECH")
        VendaOSAB.objects.create(
            pedido="SAP2",
            pdv_nome="AMAZONTECH",
            pdv_sap="1071234",
            data_abertura=agora,
        )
        cad = sincronizar_parceiros_osab()
        self.assertEqual(cad["criados"], [])
        self.assertIn("AMAZONTECH", cad["codigos_sap"])
        self.assertEqual(Parceiro.objects.get(nome="AMAZONTECH").codigo_pdv, "1071234")

    def test_codigo_real_nao_muda_mesmo_com_sap(self):
        from gestao.parceiros import sincronizar_parceiros_osab

        agora = timezone.now()
        VendaOSAB.objects.create(
            pedido="SAP3",
            pdv_nome="INOVA MG",
            pdv_sap="9999999",
            data_abertura=agora,
        )
        sincronizar_parceiros_osab()
        self.existente.refresh_from_db()
        self.assertEqual(self.existente.codigo_pdv, "1068281")

    def test_colisao_sap_mantem_placeholder(self):
        from gestao.parceiros import sincronizar_parceiros_osab

        agora = timezone.now()
        Parceiro.objects.create(codigo_pdv="OSAB-AMAZONTECH", nome="AMAZONTECH")
        VendaOSAB.objects.create(
            pedido="SAP4",
            pdv_nome="AMAZONTECH",
            pdv_sap="1068281",
            data_abertura=agora,
        )
        cad = sincronizar_parceiros_osab()
        self.assertIn("AMAZONTECH", cad["sap_colisoes"])
        self.assertEqual(
            Parceiro.objects.get(nome="AMAZONTECH").codigo_pdv, "OSAB-AMAZONTECH"
        )

    def test_preenche_especialista_e_codigo_no_existente(self):
        from gestao.parceiros import sincronizar_parceiros_osab

        agora = timezone.now()
        p = Parceiro.objects.create(codigo_pdv="OSAB-AMAZONTECH", nome="AMAZONTECH")
        VendaOSAB.objects.create(
            pedido="SAP5",
            pdv_nome="AMAZONTECH",
            pdv_sap="1071234",
            nm_gc="JESSICA TIARA",
            data_abertura=agora,
        )
        sincronizar_parceiros_osab()
        p.refresh_from_db()
        self.assertEqual(p.codigo_pdv, "1071234")
        self.assertEqual(p.especialista.first_name, "Jessica Tiara")

    def test_import_osab_grava_pdv_sap_e_atualiza_codigo(self):
        Parceiro.objects.create(codigo_pdv="OSAB-AMAZONTECH", nome="AMAZONTECH")
        hoje = timezone.localdate()
        abertura = datetime(hoje.year, hoje.month, 1, 10, 0)
        arquivo = _xlsx(
            [
                [
                    "P-SAP",
                    abertura,
                    "TT1",
                    "JOAO",
                    "AMAZONTECH",
                    abertura,
                    abertura,
                    "Concluído",
                    "500",
                    1071234.0,
                ]
            ],
            [
                "PEDIDO",
                "DT_REF",
                "MATRICULA_VENDEDOR",
                "NOME_VENDEDOR",
                "DESCRICAO",
                "DATA_ABERTURA",
                "DATA_FECHAMENTO",
                "SITUACAO",
                "VELOCIDADE",
                "PDV_SAP",
            ],
        )
        processar_osab(arquivo, "OSAB.xlsx", hoje.year, hoje.month)
        self.assertEqual(VendaOSAB.objects.get(pedido="P-SAP").pdv_sap, "1071234")
        self.assertEqual(Parceiro.objects.get(nome="AMAZONTECH").codigo_pdv, "1071234")

    def test_import_osab_grava_gerencia_no_especialista(self):
        hoje = timezone.localdate()
        abertura = datetime(hoje.year, hoje.month, 1, 10, 0)
        arquivo = _xlsx(
            [
                [
                    "P-GER",
                    abertura,
                    "TT1",
                    "JOAO",
                    "PDV GERENCIA",
                    abertura,
                    abertura,
                    "Concluído",
                    "500",
                    "MARIA FERNANDA",
                    "MG INTERIOR",
                ]
            ],
            [
                "PEDIDO",
                "DT_REF",
                "MATRICULA_VENDEDOR",
                "NOME_VENDEDOR",
                "DESCRICAO",
                "DATA_ABERTURA",
                "DATA_FECHAMENTO",
                "SITUACAO",
                "VELOCIDADE",
                "NM_GC",
                "GERENCIA",
            ],
        )
        processar_osab(arquivo, "OSAB.xlsx", hoje.year, hoje.month)
        venda = VendaOSAB.objects.get(pedido="P-GER")
        self.assertEqual(venda.gerencia, "MG INTERIOR")
        pdv = Parceiro.objects.get(nome="PDV GERENCIA")
        self.assertEqual(pdv.especialista.first_name, "Maria Fernanda")
        self.assertEqual(pdv.especialista.perfil_staff.gerencia, "MG INTERIOR")


class FpdChurnTests(TestCase):
    def setUp(self):
        self.pdv = Parceiro.objects.create(codigo_pdv="1", nome="APOLO")
        GrossMensal.objects.create(
            parceiro=self.pdv,
            anomes=int(timezone.localdate().strftime("%Y%m")),
            gross=10,
        )

    def test_fpd(self):
        mes = timezone.localdate().strftime("%m/%Y")
        arquivo = _xlsx(
            [["APOLO", mes, "Aberta", "15 a 30 Dias", "FPD"]],
            ["APELIDO", "REF_VENCTO", "SITUACAO_FATURA_MENSAL", "FAIXA", "INDICADOR"],
        )
        lote = LoteImportacao.objects.create(tipo="fpd", arquivo_nome="f.xlsx", ok=True)
        resumo = processar_fpd(arquivo, "f.xlsx", lote)
        self.assertEqual(resumo["pdvs"], 1)
        rel = RelatorioFPD.objects.get(parceiro=self.pdv)
        self.assertEqual(rel.total_abertas, 1)
        self.assertIn("Relatório FPD", rel.mensagem)
        self.assertIn("📊", rel.mensagem)
        self.assertIn("🗓️", rel.mensagem)

    def test_fpd_formato_bi(self):
        anomes = float(timezone.localdate().strftime("%Y%m"))
        arquivo = _xlsx(
            [["APOLO", anomes, "ABERTA", "0 A 15 DIAS", "FPD"]],
            ["nm_pdv_rel", "MES_VENC", "DS_SIT_FATURA", "FAIXA", "INDICADOR"],
        )
        lote = LoteImportacao.objects.create(tipo="fpd", arquivo_nome="fpd-bi.xlsx", ok=True)
        resumo = processar_fpd(arquivo, "fpd-bi.xlsx", lote)
        self.assertEqual(resumo["pdvs"], 1)
        rel = RelatorioFPD.objects.get(parceiro=self.pdv)
        self.assertEqual(rel.total_abertas, 1)
        self.assertIn("10 a 15", rel.mensagem)
        self.assertTrue(rel.detalhes.get("base"))
        self.assertIn("mes_yyyymm", rel.detalhes["meses"][0])

    def test_fpd_planilha_e_email(self):
        from gestao.fpd_format import assunto_email_fpd, html_email_fpd, planilha_fpd
        from openpyxl import load_workbook

        mes = timezone.localdate().strftime("%m/%Y")
        yyyymm = timezone.localdate().strftime("%Y%m")
        arquivo = _xlsx(
            [
                ["APOLO", mes, "Aberta", "15 a 30 Dias", "FPD", "1001"],
                ["APOLO", mes, "Paga", "0 A 15 DIAS", "FPD", "1001"],
            ],
            ["APELIDO", "REF_VENCTO", "SITUACAO_FATURA_MENSAL", "FAIXA", "INDICADOR", "cd_rede"],
        )
        lote = LoteImportacao.objects.create(tipo="fpd", arquivo_nome="f2.xlsx", ok=True)
        processar_fpd(arquivo, "f2.xlsx", lote)
        rel = RelatorioFPD.objects.get(parceiro=self.pdv)
        dados, nome = planilha_fpd(rel)
        wb = load_workbook(BytesIO(dados))
        self.assertIn("Planilha1", wb.sheetnames)
        self.assertIn("BASE_PRE_FPD_ABERTO", wb.sheetnames)
        self.assertEqual(wb["Planilha1"]["A1"].value, "MÊS FATURA")
        self.assertIn(yyyymm, str(wb["Planilha1"]["B1"].value))
        self.assertIn("FATURAS_ABERTAS", nome)
        self.assertIn("1001", assunto_email_fpd(rel))
        html = html_email_fpd(rel)
        self.assertIn("Faixas e Quantidades", html)
        self.assertIn("APOLO", html)

    def test_fpd_reimport_nao_duplica_pdv(self):
        mes = timezone.localdate().strftime("%m/%Y")
        lote1 = LoteImportacao.objects.create(tipo="fpd", arquivo_nome="f1.xlsx", ok=True)
        lote2 = LoteImportacao.objects.create(tipo="fpd", arquivo_nome="f2.xlsx", ok=True)
        for lote, abertas in ((lote1, "Aberta"), (lote2, "Paga")):
            arquivo = _xlsx(
                [["APOLO", mes, abertas, "15 a 30 Dias", "FPD"]],
                ["APELIDO", "REF_VENCTO", "SITUACAO_FATURA_MENSAL", "FAIXA", "INDICADOR"],
            )
            processar_fpd(arquivo, f"{lote.id}.xlsx", lote)
        self.assertEqual(RelatorioFPD.objects.filter(parceiro=self.pdv).count(), 1)

    def test_parse_periodo_yyyymm_float(self):
        from gestao.pipelines.fpd import _parse_periodo

        self.assertEqual(str(_parse_periodo(202608.0)), "2026-08")
        self.assertEqual(str(_parse_periodo(202608)), "2026-08")
        self.assertEqual(str(_parse_periodo("08/2026")), "2026-08")
        anomes = int(timezone.localdate().strftime("%Y%m"))
        arquivo = _xlsx(
            [["APOLO", anomes, "VOL", 0, "residencial", "preço"]],
            ["DESC_APELIDO", "ANOMES_GROSS", "TP_RETIRADA", "FLG_MEI", "NM_SEG", "DS_MOTIVO_RETIRADA"],
        )
        resumo = processar_churn(arquivo, "c.xlsx")
        self.assertEqual(resumo["pdvs"], 1)
        h = HistoricoChurn.objects.get(parceiro=self.pdv, anomes_gross=anomes)
        self.assertEqual(h.churn, 1)
        self.assertEqual(h.gross, 10)
        self.assertIn("📉", h.mensagem)
        self.assertIn("💰", h.mensagem)

    def test_resumo_capilaridade_tem_icones(self):
        from gestao.periodo import periodo_ativo
        from gestao.relatorios import resumo_geral

        ano, mes = periodo_ativo()
        txt = resumo_geral([self.pdv], ano, mes)
        self.assertIn("📊", txt)
        self.assertIn("🎯", txt)
        self.assertIn("🟢", txt)


@override_settings(
    STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {
            "BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"
        },
    }
)
class GestaoViewsTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.gestor = User.objects.create_superuser("gestor", "g@x.com", "x")
        PerfilStaff.objects.create(user=self.gestor, papel=PerfilStaff.Papel.GESTOR)

    def test_hub_exige_login(self):
        r = self.client.get(reverse("gestao_hub"))
        self.assertEqual(r.status_code, 302)

    def test_hub_gestor(self):
        self.client.force_login(self.gestor)
        r = self.client.get(reverse("gestao_hub"))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "Gestão de bases")

    def test_salvar_periodo(self):
        self.client.force_login(self.gestor)
        r = self.client.post(reverse("gestao_hub"), {"action": "periodo", "ano": 2026, "mes": 8})
        self.assertEqual(r.status_code, 302)
        self.assertEqual(periodo_ativo(), (2026, 8))

    def test_paginas_gestao(self):
        self.client.force_login(self.gestor)
        for nome in (
            "gestao_sysmap",
            "gestao_osab",
            "gestao_resultados",
            "gestao_capilaridade",
            "gestao_fpd",
            "gestao_churn",
            "gestao_configs",
            "gestao_destinatarios",
            "gestao_whatsapp",
            "gestao_envios",
        ):
            r = self.client.get(reverse(nome))
            self.assertEqual(r.status_code, 200, nome)
        cap = self.client.get(reverse("gestao_capilaridade"))
        self.assertContains(cap, "Filtros")
        self.assertContains(cap, 'name="tt"')
        self.assertContains(cap, "VENDEDOR EXTERNO")
        self.assertContains(cap, "Meus parceiros")
        self.assertContains(cap, "Outros especialistas")
        sysmap = self.client.get(reverse("gestao_sysmap"))
        self.assertContains(sysmap, "Sysmap / Supply")
        self.assertContains(sysmap, "Importar Sysmap")
        cap = self.client.get(reverse("gestao_capilaridade"))
        self.assertContains(cap, "Sysmap")

    @override_settings(
        EVOLUTION_API_URL="https://evo.test",
        EVOLUTION_API_KEY="evo.key",
        EVOLUTION_INSTANCE_NAME="nio_gc_tickets",
    )
    def test_tarefas_tem_enviar_todos(self):
        self.client.force_login(self.gestor)
        r = self.client.get(reverse("gestao_tarefas"))
        self.assertContains(r, "Enviar todos (WhatsApp)")

    @override_settings(
        EVOLUTION_API_URL="https://evo.test",
        EVOLUTION_API_KEY="evo.key",
        EVOLUTION_INSTANCE_NAME="nio_gc_tickets",
    )
    def test_tarefas_enviar_todos_post(self):
        from unittest.mock import patch

        from gestao.messaging.envio import ResumoEnvio

        self.client.force_login(self.gestor)
        with patch(
            "gestao.views.enviar_tarefas_todos",
            return_value=ResumoEnvio(enviados=1, detalhes=["ok"]),
        ) as mock_env:
            r = self.client.post(reverse("gestao_tarefas"), {"action": "enviar_todos"})
        self.assertEqual(r.status_code, 302)
        mock_env.assert_called_once()

    def test_cadastrar_parceiros_da_osab(self):
        self.client.force_login(self.gestor)
        VendaOSAB.objects.create(
            pedido="Z1", pdv_nome="PDV NOVO OSAB", data_abertura=timezone.now()
        )
        osab = self.client.get(reverse("gestao_osab"))
        self.assertContains(osab, "Cadastrar PDVs que faltam")
        self.assertContains(osab, "PDV NOVO OSAB")
        r = self.client.post(
            reverse("gestao_osab"), {"action": "cadastrar_parceiros"}
        )
        self.assertEqual(r.status_code, 302)
        self.assertTrue(Parceiro.objects.filter(nome="PDV NOVO OSAB").exists())

    def test_osab_lista_um_relatorio_por_pdv(self):
        from gestao.models import HistoricoOSAB

        pdv = Parceiro.objects.create(
            codigo_pdv="dup1", nome="ALLVO TELECOM", especialista=self.gestor
        )
        HistoricoOSAB.objects.create(
            parceiro=pdv,
            descricao_pdv="ALLVO TELECOM",
            status="Sem metas",
            mensagem="snapshot ontem",
        )
        HistoricoOSAB.objects.create(
            parceiro=pdv,
            descricao_pdv="ALLVO TELECOM",
            status="Ok",
            mensagem="Relatório financeiro e desempenho",
        )
        self.client.force_login(self.gestor)
        html = self.client.get(reverse("gestao_osab")).content.decode()
        self.assertEqual(html.count('class="mask-card"'), 1)
        self.assertIn("Relatório financeiro e desempenho", html)
        self.assertNotIn("snapshot ontem", html)

    def test_paginas_especialista_sem_whatsapp_qr(self):
        User = get_user_model()
        spec = User.objects.create_user("specg", "sg@x.com", "x", is_staff=True)
        PerfilStaff.objects.create(user=spec, papel=PerfilStaff.Papel.ESPECIALISTA)
        self.client.force_login(spec)
        for nome in (
            "gestao_hub",
            "gestao_osab",
            "gestao_resultados",
            "gestao_capilaridade",
            "gestao_fpd",
            "gestao_churn",
            "gestao_comissionamento",
            "gestao_tarefas",
            "gestao_venda_indevida",
            "gestao_recompra",
            "gestao_configs",
            "gestao_envios",
        ):
            r = self.client.get(reverse(nome))
            self.assertEqual(r.status_code, 200, nome)
            self.assertContains(r, "Capilaridade")
            self.assertContains(r, "Meus parceiros")
        self.assertEqual(self.client.get(reverse("gestao_whatsapp")).status_code, 404)
        self.assertEqual(self.client.get(reverse("gestao_destinatarios")).status_code, 404)

    def test_especialista_importa_bases(self):
        from gestao.periodo import periodo_ativo

        User = get_user_model()
        spec = User.objects.create_user("specimp", "si@x.com", "x", is_staff=True)
        PerfilStaff.objects.create(user=spec, papel=PerfilStaff.Papel.ESPECIALISTA)
        self.client.force_login(spec)
        self.assertContains(self.client.get(reverse("gestao_osab")), "Importar OSAB")
        self.assertContains(self.client.get(reverse("gestao_sysmap")), "Importar Sysmap")
        self.assertContains(self.client.get(reverse("gestao_fpd")), "Importar FPD")
        self.assertContains(self.client.get(reverse("gestao_tarefas")), "Importar tarefas")
        self.assertContains(
            self.client.get(reverse("gestao_configs")), "Importar acompanhamento semanal"
        )
        r = self.client.post(
            reverse("gestao_hub"), {"action": "periodo", "ano": 2026, "mes": 7}
        )
        self.assertEqual(r.status_code, 302)
        self.assertEqual(periodo_ativo(), (2026, 7))
        bloqueado = self.client.post(
            reverse("gestao_configs"), {"action": "salvar_politica"}
        )
        self.assertEqual(bloqueado.status_code, 302)


class SyncWAClientTests(TestCase):
    def test_normalizar_numero_br(self):
        from gestao.messaging.syncwa import normalizar_destino, numero_para_evolution

        self.assertEqual(normalizar_destino("31999999999"), "5531999999999@s.whatsapp.net")
        self.assertEqual(normalizar_destino("5531999999999"), "5531999999999@s.whatsapp.net")
        self.assertEqual(normalizar_destino("120363@g.us"), "120363@g.us")
        self.assertEqual(numero_para_evolution("5531999999999@s.whatsapp.net"), "5531999999999")
        self.assertEqual(numero_para_evolution("120363418335765186@g.us"), "120363418335765186@g.us")

    @override_settings(
        EVOLUTION_API_URL="https://evo.test",
        EVOLUTION_API_KEY="evo.key",
        EVOLUTION_INSTANCE_NAME="nio_gc_tickets",
        N8N_OUTBOUND_WEBHOOK_URL="",
        SYNCWA_MODO_TESTE=False,
    )
    def test_enviar_documento_usa_base64_puro(self):
        import base64
        from unittest.mock import MagicMock, patch

        from gestao.messaging.syncwa import enviar_documento

        fake = MagicMock()
        fake.status_code = 200
        fake.json.return_value = {"key": {"id": "doc-1"}}
        with patch("gestao.messaging.syncwa.requests.post", return_value=fake) as post:
            result = enviar_documento(
                "5531999999999", conteudo=b"abc", file_name="Tarefas VISION.xlsx"
            )
        self.assertTrue(result.ok)
        payload = post.call_args.kwargs["json"]
        self.assertEqual(payload["media"], base64.b64encode(b"abc").decode("ascii"))
        self.assertFalse(payload["media"].startswith("data:"))
        self.assertEqual(payload["fileName"], "Tarefas_VISION.xlsx")
        self.assertEqual(
            payload["mimetype"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("/message/sendMedia/nio_gc_tickets", post.call_args.args[0])

    @override_settings(
        EVOLUTION_API_URL="https://evo.test",
        EVOLUTION_API_KEY="evo.key",
        EVOLUTION_INSTANCE_NAME="nio_gc_tickets",
        N8N_OUTBOUND_WEBHOOK_URL="",
        SYNCWA_MODO_TESTE=False,
    )
    def test_enviar_documento_tenta_data_uri_se_base64_rejeitado(self):
        from unittest.mock import MagicMock, patch

        from gestao.messaging.syncwa import enviar_documento

        ruim = MagicMock()
        ruim.status_code = 400
        ruim.json.return_value = {
            "status": 400,
            "error": "Bad Request",
            "response": {"message": ["Owned media must be a url or base64"]},
        }
        bom = MagicMock()
        bom.status_code = 200
        bom.json.return_value = {"key": {"id": "doc-2"}}
        with patch("gestao.messaging.syncwa.requests.post", side_effect=[ruim, bom]) as post:
            result = enviar_documento("5531999999999", conteudo=b"xyz", file_name="a.xlsx")
        self.assertTrue(result.ok)
        self.assertEqual(post.call_count, 2)
        self.assertTrue(post.call_args_list[1].kwargs["json"]["media"].startswith("data:"))

    @override_settings(
        EVOLUTION_API_URL="https://evo.test",
        EVOLUTION_API_KEY="evo.key",
        EVOLUTION_INSTANCE_NAME="nio_gc_tickets",
        N8N_OUTBOUND_WEBHOOK_URL="",
        SYNCWA_MODO_TESTE=False,
    )
    def test_enviar_texto_registra_ok(self):
        from unittest.mock import MagicMock, patch

        from gestao.messaging.syncwa import enviar_texto

        fake = MagicMock()
        fake.status_code = 200
        fake.json.return_value = {"key": {"id": "ml-1"}}
        with patch("gestao.messaging.syncwa.requests.post", return_value=fake) as post:
            result = enviar_texto("5531999999999", "oi")
        self.assertTrue(result.ok)
        self.assertEqual(result.message_log_id, "ml-1")
        post.assert_called_once()
        kwargs = post.call_args.kwargs
        self.assertEqual(kwargs["json"]["number"], "5531999999999")
        self.assertEqual(kwargs["headers"]["apikey"], "evo.key")
        self.assertIn("/message/sendText/nio_gc_tickets", post.call_args.args[0])

    @override_settings(
        EVOLUTION_API_URL="https://evo.test",
        EVOLUTION_API_KEY="evo.key",
        EVOLUTION_INSTANCE_NAME="nio_gc_tickets",
        N8N_OUTBOUND_WEBHOOK_URL="",
        SYNCWA_MODO_TESTE=True,
        SYNCWA_TEST_JID="5531888888888",
    )
    def test_modo_teste_redireciona(self):
        from unittest.mock import MagicMock, patch

        from gestao.messaging.syncwa import enviar_texto

        fake = MagicMock()
        fake.status_code = 200
        fake.json.return_value = {"key": {"id": "ml-2"}}
        with patch("gestao.messaging.syncwa.requests.post", return_value=fake) as post:
            result = enviar_texto("120363xxx@g.us", "teste")
        self.assertTrue(result.ok)
        self.assertEqual(post.call_args.kwargs["json"]["number"], "5531888888888")


@override_settings(
    STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {
            "BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"
        },
    },
    EVOLUTION_API_URL="https://evo.test",
    EVOLUTION_API_KEY="evo.key",
    EVOLUTION_INSTANCE_NAME="nio_gc_tickets",
    N8N_OUTBOUND_WEBHOOK_URL="",
)
class WhatsAppPareamentoTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.gestor = User.objects.create_superuser("gwa", "gwa@x.com", "x")
        PerfilStaff.objects.create(user=self.gestor, papel=PerfilStaff.Papel.GESTOR)
        self.client.force_login(self.gestor)

    def test_pagina_gestor(self):
        r = self.client.get(reverse("gestao_whatsapp"))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "Gerar QR Code")
        self.assertContains(r, "nio_gc_tickets")

    def test_especialista_nao_acessa(self):
        User = get_user_model()
        spec = User.objects.create_user("specwa", "sw@x.com", "x", is_staff=True)
        PerfilStaff.objects.create(user=spec, papel=PerfilStaff.Papel.ESPECIALISTA)
        self.client.force_login(spec)
        r = self.client.get(reverse("gestao_whatsapp"))
        self.assertEqual(r.status_code, 404)

    @override_settings(EVOLUTION_API_URL="", EVOLUTION_API_KEY="")
    def test_status_sem_config(self):
        r = self.client.get(reverse("gestao_whatsapp_status"))
        self.assertEqual(r.status_code, 503)
        self.assertEqual(r.json()["state"], "unconfigured")

    def test_status_conectado(self):
        from unittest.mock import patch

        with patch(
            "gestao.views_whatsapp.EvolutionConnectionService.get_status",
            return_value={
                "instanceName": "nio_gc_tickets",
                "state": "open",
                "connected": True,
                "n8nConfigured": False,
                "evolutionConfigured": True,
            },
        ):
            r = self.client.get(reverse("gestao_whatsapp_status"))
        self.assertEqual(r.status_code, 200)
        self.assertTrue(r.json()["connected"])
        self.assertEqual(r.json()["state"], "open")

    def test_qrcode(self):
        from unittest.mock import patch

        with patch(
            "gestao.views_whatsapp.EvolutionConnectionService.get_qrcode",
            return_value={
                "instanceName": "nio_gc_tickets",
                "base64": "data:image/png;base64,AAA",
                "count": 1,
            },
        ):
            r = self.client.get(reverse("gestao_whatsapp_qrcode"))
        self.assertEqual(r.status_code, 200)
        self.assertTrue(r.json()["base64"].startswith("data:image/png;base64,"))

    def test_desconectar(self):
        from unittest.mock import patch

        with patch(
            "gestao.views_whatsapp.EvolutionConnectionService.disconnect",
            return_value={
                "success": True,
                "instanceName": "nio_gc_tickets",
                "message": "Instância desconectada com sucesso.",
                "status": {"state": "close", "connected": False},
            },
        ):
            r = self.client.post(reverse("gestao_whatsapp_disconnect"))
        self.assertEqual(r.status_code, 200)
        self.assertTrue(r.json()["success"])


@override_settings(
    STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {
            "BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"
        },
    },
    EVOLUTION_API_URL="https://evo.test",
    EVOLUTION_API_KEY="evo.key",
    EVOLUTION_INSTANCE_NAME="nio_gc_tickets",
    N8N_OUTBOUND_WEBHOOK_URL="",
    SYNCWA_MODO_TESTE=False,
)
class DestinatarioEnvioTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.gestor = User.objects.create_superuser("g2", "g2@x.com", "x")
        PerfilStaff.objects.create(user=self.gestor, papel=PerfilStaff.Papel.GESTOR)
        self.pdv = Parceiro.objects.create(codigo_pdv="99", nome="INOVA TESTE")
        self.client.force_login(self.gestor)

    def test_cria_destinatario(self):
        from gestao.models import Destinatario

        r = self.client.post(
            reverse("gestao_destinatarios"),
            {
                "parceiro": self.pdv.id,
                "nome": "Grupo Inova",
                "jid": "120363abc@g.us",
                "tipo": Destinatario.TipoDestino.GRUPO,
                "prioridade": 10,
                "ativo": "on",
                "envio_capilaridade": "on",
                "envio_osab": "on",
            },
        )
        self.assertEqual(r.status_code, 302)
        dest = Destinatario.objects.get(nome="Grupo Inova")
        self.assertTrue(dest.envio_capilaridade)
        self.assertFalse(dest.envio_fpd)

    def test_sincroniza_whatsapp_do_especialista_em_massa(self):
        from gestao.models import Destinatario

        User = get_user_model()
        spec = User.objects.create_user(
            "ana.gc", "ana@x.com", "x", is_staff=True, first_name="Ana GC"
        )
        PerfilStaff.objects.create(
            user=spec, papel=PerfilStaff.Papel.ESPECIALISTA, whatsapp="31988887777"
        )
        self.pdv.especialista = spec
        self.pdv.save(update_fields=["especialista"])
        grupo = Destinatario.objects.create(
            parceiro=self.pdv,
            nome="Grupo PDV",
            jid="120363grupo@g.us",
            tipo=Destinatario.TipoDestino.GRUPO,
            envio_capilaridade=True,
        )
        r = self.client.post(
            reverse("gestao_destinatarios"),
            {"action": "sincronizar_especialistas"},
        )
        self.assertEqual(r.status_code, 302)
        dest = Destinatario.objects.get(
            parceiro=self.pdv, tipo=Destinatario.TipoDestino.INDIVIDUAL
        )
        self.assertEqual(dest.jid, "5531988887777")
        self.assertTrue(dest.nome.startswith("Especialista:"))
        self.assertFalse(dest.envio_osab)
        self.assertTrue(dest.envio_tarefas)
        self.assertTrue(Destinatario.objects.filter(pk=grupo.pk).exists())
        r2 = self.client.post(
            reverse("gestao_destinatarios"),
            {"action": "sincronizar_especialistas"},
        )
        self.assertEqual(r2.status_code, 302)
        self.assertEqual(
            Destinatario.objects.filter(
                parceiro=self.pdv, tipo=Destinatario.TipoDestino.INDIVIDUAL
            ).count(),
            1,
        )
        spec.perfil_staff.whatsapp = "5531911112222"
        spec.perfil_staff.save(update_fields=["whatsapp"])
        self.client.post(
            reverse("gestao_destinatarios"),
            {"action": "sincronizar_especialistas"},
        )
        dest.refresh_from_db()
        self.assertEqual(dest.jid, "5531911112222")

    def test_sincroniza_nao_cria_dest_do_admin_e_remove_o_existente(self):
        from gestao.models import Destinatario

        self.pdv.especialista = self.gestor
        self.pdv.save(update_fields=["especialista"])
        self.gestor.perfil_staff.whatsapp = "21979630377"
        self.gestor.perfil_staff.save(update_fields=["whatsapp"])
        Destinatario.objects.create(
            parceiro=self.pdv,
            nome="Especialista: Rogério Pereira Pacheco",
            jid="5521979630377",
            tipo=Destinatario.TipoDestino.INDIVIDUAL,
            envio_osab=True,
        )
        grupo = Destinatario.objects.create(
            parceiro=self.pdv,
            nome="Grupo PDV",
            jid="120363grupo@g.us",
            tipo=Destinatario.TipoDestino.GRUPO,
        )
        r = self.client.post(
            reverse("gestao_destinatarios"),
            {"action": "sincronizar_especialistas"},
        )
        self.assertEqual(r.status_code, 302)
        self.assertFalse(
            Destinatario.objects.filter(
                parceiro=self.pdv, tipo=Destinatario.TipoDestino.INDIVIDUAL
            ).exists()
        )
        self.assertTrue(Destinatario.objects.filter(pk=grupo.pk).exists())

    def test_destinos_osab_empresario_e_especialista(self):
        from gestao.messaging.envio import destinos_para_envio
        from gestao.models import Destinatario
        from tickets.models import ContatoParceiro

        User = get_user_model()
        spec = User.objects.create_user(
            "specosab", "so@x.com", "x", is_staff=True, first_name="Ana"
        )
        PerfilStaff.objects.create(
            user=spec, papel=PerfilStaff.Papel.ESPECIALISTA, whatsapp="31988887777"
        )
        self.pdv.especialista = spec
        self.pdv.save(update_fields=["especialista"])
        ContatoParceiro.objects.create(
            parceiro=self.pdv,
            nome="Dono",
            telefone="21987654321",
            cargo=ContatoParceiro.Cargo.EMPRESARIO,
        )
        ContatoParceiro.objects.create(
            parceiro=self.pdv,
            nome="BO",
            telefone="21911112222",
            cargo=ContatoParceiro.Cargo.BACKOFFICE,
        )
        Destinatario.objects.create(
            parceiro=self.pdv,
            nome="Grupo PDV",
            jid="120363grupo@g.us",
            tipo=Destinatario.TipoDestino.GRUPO,
            envio_osab=True,
        )
        destinos = destinos_para_envio(self.gestor, "envio_osab", self.pdv)
        jids = [d.jid for d in destinos]
        self.assertEqual(jids, ["5521987654321", "5531988887777"])
        self.assertNotIn("120363grupo@g.us", jids)

        so_empresario = destinos_para_envio(spec, "envio_osab", self.pdv)
        self.assertEqual([d.jid for d in so_empresario], ["5521987654321"])

    def test_destinos_osab_pdv_do_admin_nao_inclui_o_proprio_numero(self):
        from gestao.messaging.envio import destinos_para_envio
        from tickets.models import ContatoParceiro

        self.pdv.especialista = self.gestor
        self.pdv.save(update_fields=["especialista"])
        self.gestor.perfil_staff.whatsapp = "21979630377"
        self.gestor.perfil_staff.save(update_fields=["whatsapp"])
        ContatoParceiro.objects.create(
            parceiro=self.pdv,
            nome="Empresário do PDV",
            telefone="11999998888",
            cargo="Empresário",
        )
        destinos = destinos_para_envio(self.gestor, "envio_osab", self.pdv)
        self.assertEqual([d.jid for d in destinos], ["5511999998888"])

    def test_destinos_osab_sem_empresario_usa_grupo(self):
        from gestao.messaging.envio import destinos_para_envio
        from gestao.models import Destinatario

        self.pdv.especialista = self.gestor
        self.pdv.save(update_fields=["especialista"])
        Destinatario.objects.create(
            parceiro=self.pdv,
            nome="GERÊNCIA VISION",
            jid="120363424950507669@g.us",
            tipo=Destinatario.TipoDestino.GRUPO,
            envio_osab=True,
        )
        destinos = destinos_para_envio(self.gestor, "envio_osab", self.pdv)
        self.assertEqual([d.jid for d in destinos], ["120363424950507669@g.us"])
        self.assertEqual(destinos[0].nome, "GERÊNCIA VISION")

    def test_destinos_comissionamento_empresario(self):
        from gestao.messaging.envio import destinos_para_envio
        from tickets.models import ContatoParceiro

        ContatoParceiro.objects.create(
            parceiro=self.pdv,
            nome="Empresário 1",
            telefone="31999990001",
            cargo=ContatoParceiro.Cargo.EMPRESARIO,
        )
        ContatoParceiro.objects.create(
            parceiro=self.pdv,
            nome="Backoffice",
            telefone="31999990002",
            cargo=ContatoParceiro.Cargo.BACKOFFICE,
        )
        destinos = destinos_para_envio(self.gestor, "envio_comissionamento", self.pdv)
        self.assertEqual([d.jid for d in destinos], ["5531999990001"])
        self.assertEqual(destinos[0].nome, "Empresário 1")

    def test_destinos_comissionamento_fallback_telefone_ou_destinatario(self):
        from gestao.messaging.envio import destinos_para_envio
        from gestao.models import Destinatario

        # 1. Sem contatos empresário, usa o telefone do parceiro
        self.pdv.telefone = "31988887777"
        self.pdv.save(update_fields=["telefone"])
        destinos = destinos_para_envio(self.gestor, "envio_comissionamento", self.pdv)
        self.assertEqual([d.jid for d in destinos], ["5531988887777"])

        # 2. Sem telefone no parceiro, fallback para Destinatario legado
        self.pdv.telefone = ""
        self.pdv.save(update_fields=["telefone"])
        Destinatario.objects.create(
            parceiro=self.pdv,
            nome="Grupo Comissao",
            jid="120363comissao@g.us",
            tipo=Destinatario.TipoDestino.GRUPO,
            envio_comissionamento=True,
        )
        destinos2 = destinos_para_envio(self.gestor, "envio_comissionamento", self.pdv)
        self.assertEqual([d.jid for d in destinos2], ["120363comissao@g.us"])

    def test_destinos_comissionamento_diretora_e_empresaria(self):
        from gestao.messaging.envio import destinos_para_envio
        from tickets.models import ContatoParceiro

        ContatoParceiro.objects.create(
            parceiro=self.pdv,
            nome="Anacleto",
            telefone="33984590017",
            cargo="Diretor",
        )
        ContatoParceiro.objects.create(
            parceiro=self.pdv,
            nome="Janaina",
            telefone="33988310607",
            cargo="Diretora",
        )
        destinos = destinos_para_envio(self.gestor, "envio_comissionamento", self.pdv)
        self.assertEqual(
            [d.jid for d in destinos],
            ["5533984590017", "5533988310607"],
        )

    def test_enviar_capilaridade_registra_log(self):
        from unittest.mock import MagicMock, patch

        from gestao.models import Destinatario, EnvioWhatsApp

        Destinatario.objects.create(
            parceiro=self.pdv,
            nome="Contato",
            jid="5531999999999",
            envio_capilaridade=True,
        )
        fake = MagicMock()
        fake.status_code = 200
        fake.json.return_value = {"key": {"id": "ml-9"}}
        with patch("gestao.messaging.syncwa.requests.post", return_value=fake):
            r = self.client.post(
                reverse("gestao_capilaridade"),
                {"action": "enviar_pdv", "parceiro": self.pdv.id, "escopo": "outros"},
            )
        self.assertEqual(r.status_code, 302)
        logs = EnvioWhatsApp.objects.filter(tipo=EnvioWhatsApp.Tipo.CAPILARIDADE)
        self.assertTrue(logs.exists())
        self.assertTrue(all(l.status == EnvioWhatsApp.Status.ENVIADO for l in logs))
        self.assertTrue(any(l.syncwa_message_id == "ml-9" for l in logs))

    def test_destinos_especialista_usa_whatsapp_do_perfil(self):
        from gestao.messaging.envio import destinos_para_envio
        from gestao.models import Destinatario

        Destinatario.objects.create(
            parceiro=self.pdv,
            nome="Grupo PDV",
            jid="120363grupo@g.us",
            envio_capilaridade=True,
        )
        User = get_user_model()
        spec = User.objects.create_user("specenvio", "se@x.com", "x", is_staff=True, first_name="Ana")
        perfil = PerfilStaff.objects.create(
            user=spec, papel=PerfilStaff.Papel.ESPECIALISTA, whatsapp="5531988887777"
        )
        destinos = destinos_para_envio(spec, "envio_capilaridade", self.pdv)
        self.assertEqual(len(destinos), 1)
        self.assertEqual(destinos[0].jid, "5531988887777")
        self.assertIsNone(destinos[0].destinatario)

        gestor_dest = destinos_para_envio(self.gestor, "envio_capilaridade", self.pdv)
        self.assertEqual(len(gestor_dest), 1)
        self.assertEqual(gestor_dest[0].jid, "120363grupo@g.us")

        perfil.whatsapp = ""
        perfil.save(update_fields=["whatsapp"])
        self.assertEqual(destinos_para_envio(spec, "envio_capilaridade", self.pdv), [])

    def test_emails_so_gestor(self):
        from gestao.messaging.envio import emails_para_envio
        from gestao.models import Destinatario

        Destinatario.objects.create(
            parceiro=self.pdv,
            nome="Grupo PDV",
            jid="120363grupo@g.us",
            email="grupo@x.com",
            email_capilaridade=True,
            envio_capilaridade=True,
        )
        User = get_user_model()
        spec = User.objects.create_user("specmailg", "smg@x.com", "x", is_staff=True)
        PerfilStaff.objects.create(user=spec, papel=PerfilStaff.Papel.ESPECIALISTA)
        self.assertEqual(emails_para_envio(spec, "envio_capilaridade", self.pdv), [])
        self.assertEqual(
            emails_para_envio(self.gestor, "envio_capilaridade", self.pdv),
            ["grupo@x.com"],
        )

    def test_email_fpd_vai_para_especialista_pdv(self):
        from gestao.messaging.envio import emails_fpd_especialista

        User = get_user_model()
        spec = User.objects.create_user(
            "specfpd", "rogerio.pacheco@niointernet.com.br", "x", is_staff=True
        )
        PerfilStaff.objects.create(user=spec, papel=PerfilStaff.Papel.ESPECIALISTA)
        self.pdv.especialista = spec
        self.pdv.save(update_fields=["especialista"])
        self.assertEqual(
            emails_fpd_especialista(self.pdv),
            ["rogerio.pacheco@niointernet.com.br"],
        )
        self.pdv.especialista = None
        self.pdv.save(update_fields=["especialista"])
        self.assertEqual(emails_fpd_especialista(self.pdv), [])

    def test_especialista_envia_capilaridade_para_si(self):
        from unittest.mock import MagicMock, patch

        from gestao.models import Destinatario, EnvioWhatsApp

        Destinatario.objects.create(
            parceiro=self.pdv,
            nome="Grupo PDV",
            jid="120363grupo@g.us",
            envio_capilaridade=True,
        )
        User = get_user_model()
        spec = User.objects.create_user("specmask", "sm@x.com", "x", is_staff=True)
        PerfilStaff.objects.create(
            user=spec, papel=PerfilStaff.Papel.ESPECIALISTA, whatsapp="5531911112222"
        )
        self.pdv.especialista = spec
        self.pdv.save(update_fields=["especialista"])
        self.client.force_login(spec)
        fake = MagicMock()
        fake.status_code = 200
        fake.json.return_value = {"key": {"id": "ml-esp"}}
        with patch("gestao.messaging.syncwa.requests.post", return_value=fake) as post:
            r = self.client.post(
                reverse("gestao_capilaridade"),
                {"action": "enviar_pdv", "parceiro": self.pdv.id},
            )
        self.assertEqual(r.status_code, 302)
        logs = EnvioWhatsApp.objects.filter(tipo=EnvioWhatsApp.Tipo.CAPILARIDADE)
        self.assertTrue(logs.exists())
        self.assertEqual(logs.first().criado_por, spec)
        for log in logs:
            self.assertIn("5531911112222", log.destino_jid)
            self.assertNotIn("120363grupo", log.destino_jid)
        for call in post.call_args_list:
            self.assertNotIn("120363grupo", call.kwargs["json"]["number"])

    def test_enviar_teste_via_envios(self):
        from unittest.mock import MagicMock, patch

        from gestao.models import EnvioWhatsApp

        fake = MagicMock()
        fake.status_code = 200
        fake.json.return_value = {"key": {"id": "ml-t"}}
        with patch("gestao.messaging.syncwa.requests.post", return_value=fake):
            with override_settings(SYNCWA_TEST_JID="5531777777777"):
                r = self.client.post(reverse("gestao_envios"), {"action": "teste"})
        self.assertEqual(r.status_code, 302)
        self.assertTrue(EnvioWhatsApp.objects.filter(tipo=EnvioWhatsApp.Tipo.TESTE).exists())


class ComissionamentoTests(TestCase):
    def setUp(self):
        self.pdv = Parceiro.objects.create(
            codigo_pdv="31",
            nome="INOVA MG",
            razao_social="LUISA SERVICOS DE TELEFONIA MOVEL LTDA",
        )

    def _ciclo_xlsx(self):
        wb = Workbook()
        ws_p = wb.active
        ws_p.title = "PEDIDO"
        ws_p.append(
            [
                "DOCUMENTO DE COMPRAS",
                "ITEM",
                "VALOR",
                "FORNECEDOR",
                "RAZAO SOCIAL",
                "CNPJ",
                "CANAL",
                "CENTRO",
                "CICLO",
            ]
        )
        ws_p.append(
            [
                "450001",
                "10",
                "100,50",
                "HANA1",
                "LUISA SERVICOS DE TELEFONIA MOVEL LTDA",
                "123",
                "VAREJO",
                "MG",
                "202608",
            ]
        )
        ws_p.append(
            [
                "450002",
                "20",
                "50,00",
                "HANA2",
                "OUTRA EMPRESA LTDA",
                "999",
                "VAREJO",
                "MG",
                "202608",
            ]
        )
        ws_l = wb.create_sheet("LINHA_A_LINHA")
        ws_l.append(["RAZAO SOCIAL", "SUB_EVENTO", "COMISSAO"])
        ws_l.append(["LUISA SERVICOS DE TELEFONIA MOVEL LTDA", "ATIVACAO", "100,50"])
        ws_l.append(["OUTRA EMPRESA LTDA", "ATIVACAO", "50,00"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "ciclo.xlsx"
        return buf

    def test_processar_filtra_por_razao(self):
        from gestao.models import LoteImportacao, RelatorioComissionamento
        from gestao.pipelines.comissionamento import processar_comissionamento

        lote = LoteImportacao.objects.create(
            tipo=LoteImportacao.Tipo.COMISSIONAMENTO,
            arquivo_nome="ciclo.xlsx",
            ok=True,
        )
        resumo = processar_comissionamento(self._ciclo_xlsx(), "ciclo.xlsx", lote)
        self.assertEqual(resumo["pdvs"], 1)
        rel = RelatorioComissionamento.objects.get()
        self.assertEqual(rel.parceiro_id, self.pdv.id)
        self.assertEqual(rel.qtd_pedido, 1)
        self.assertEqual(rel.qtd_linha, 1)
        self.assertTrue(rel.arquivo)
        self.assertIn("Comissionamento", rel.mensagem)
        self.assertIn("Orientação para envio do email:", rel.mensagem)
        self.assertIn("recebimentonfes@niointernet.com.br", rel.mensagem)
        self.assertIn("rogerio.pacheco@niointernet.com.br", rel.mensagem)
        self.assertIn("PP-GestaodosParceiros@niointernet.com.br", rel.mensagem)
        self.assertIn("No corpo do email retirar assinatura e não escrever nada no corpo do email", rel.mensagem)
        self.assertIn("E o assunto do email deverá ser", rel.mensagem)

        import openpyxl
        rel.arquivo.open("rb")
        wb = openpyxl.load_workbook(rel.arquivo)
        self.assertIn("PEDIDO", wb.sheetnames)
        self.assertIn("LINHA_A_LINHA", wb.sheetnames)
        ws_pedido = wb["PEDIDO"]
        self.assertEqual(ws_pedido.cell(1, 1).fill.start_color.rgb, "001A365D")
        self.assertEqual(ws_pedido.cell(1, 1).font.color.rgb, "00FFFFFF")
        self.assertTrue(ws_pedido.views.sheetView[0].showGridLines)

    def test_botao_baixar_planilha_presente_na_view(self):
        from django.contrib.auth import get_user_model
        from gestao.models import LoteImportacao, RelatorioComissionamento
        from gestao.pipelines.comissionamento import processar_comissionamento
        from tickets.models import PerfilStaff

        User = get_user_model()
        user = User.objects.create_superuser("admin_comiss", "admin@teste.com", "senha123")
        PerfilStaff.objects.create(user=user, papel=PerfilStaff.Papel.GESTOR)
        self.pdv.especialista = user
        self.pdv.save(update_fields=["especialista"])
        self.client.force_login(user)

        lote = LoteImportacao.objects.create(
            tipo=LoteImportacao.Tipo.COMISSIONAMENTO,
            arquivo_nome="ciclo.xlsx",
            ok=True,
        )
        processar_comissionamento(self._ciclo_xlsx(), "ciclo.xlsx", lote)
        resp = self.client.get("/gestao/comissionamento/")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "📥 Baixar Planilha")

    def test_exige_razoes_configuradas(self):
        from gestao.models import LoteImportacao
        from gestao.pipelines.comissionamento import processar_comissionamento

        self.pdv.razao_social = ""
        self.pdv.save(update_fields=["razao_social"])
        lote = LoteImportacao.objects.create(
            tipo=LoteImportacao.Tipo.COMISSIONAMENTO,
            arquivo_nome="ciclo.xlsx",
            ok=True,
        )
        with self.assertRaises(ValueError):
            processar_comissionamento(self._ciclo_xlsx(), "ciclo.xlsx", lote)


class TarefasTests(TestCase):
    def setUp(self):
        self.pdv = Parceiro.objects.create(codigo_pdv="31", nome="INOVA MG")

    def _xlsx_abertas(self, data_agendamento):
        from datetime import date as d

        if isinstance(data_agendamento, d):
            data_agendamento = data_agendamento.isoformat()
        wb = Workbook()
        ws = wb.active
        ws.append(
            [
                "sg_uf",
                "nm_municipio",
                "INDICADOR",
                "nm_pdv_rel",
                "DT_AGENDAMENTO",
                "nr_ordem",
            ]
        )
        ws.append(["MG", "BH", "TAREFAS ABERTAS", "INOVA MG", data_agendamento, "1"])
        ws.append(["MG", "Uberlândia", "TAREFAS ABERTAS", "INOVA MG", data_agendamento, "2"])
        ws.append(["SP", "SP", "TAREFAS ABERTAS", "OUTRO", data_agendamento, "3"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "tarefas.xlsx"
        return buf

    def test_abertas_por_pdv(self):
        from gestao.models import LoteImportacao, RelatorioTarefa
        from gestao.periodo import hoje
        from gestao.pipelines.tarefas import processar_tarefas

        lote = LoteImportacao.objects.create(
            tipo=LoteImportacao.Tipo.TAREFAS, arquivo_nome="t.xlsx", ok=True
        )
        resumo = processar_tarefas(self._xlsx_abertas(hoje()), "t.xlsx", lote)
        self.assertEqual(resumo["modo"], "abertas")
        self.assertEqual(resumo["relatorios"], 2)
        self.assertEqual(resumo["total_hoje"], 3)
        rel = RelatorioTarefa.objects.get(parceiro=self.pdv)
        self.assertEqual(rel.total, 2)
        self.assertIn("Resumo de Tarefas", rel.mensagem)
        self.assertNotIn("(MG) -", rel.mensagem)
        self.assertIn("BH (MG)", rel.mensagem)
        outro = RelatorioTarefa.objects.get(pdv_nome="OUTRO")
        self.assertEqual(outro.total, 1)
        self.assertIn("SP (SP)", outro.mensagem)
        self.assertIsNone(outro.parceiro_id)

    def test_abertas_inclui_todas_as_ufs(self):
        from gestao.models import LoteImportacao, RelatorioTarefa
        from gestao.periodo import hoje
        from gestao.pipelines.tarefas import processar_tarefas

        lote = LoteImportacao.objects.create(
            tipo=LoteImportacao.Tipo.TAREFAS, arquivo_nome="t.xlsx", ok=True
        )
        processar_tarefas(self._xlsx_abertas(hoje()), "t.xlsx", lote)
        totais = {
            r.pdv_nome: r.total for r in RelatorioTarefa.objects.all()
        }
        self.assertEqual(totais["INOVA MG"], 2)
        self.assertEqual(totais["OUTRO"], 1)

    def test_fechadas_consolidado(self):
        from gestao.models import LoteImportacao, RelatorioTarefa
        from gestao.periodo import hoje
        from gestao.pipelines.tarefas import processar_tarefas

        wb = Workbook()
        ws = wb.active
        ws.append(["sg_uf", "nm_municipio", "INDICADOR", "dt_fim_execucao_real"])
        ws.append(["MG", "BH", "TAREFAS FECHADAS", hoje().isoformat()])
        ws.append(["RJ", "Rio de Janeiro", "TAREFAS FECHADAS", hoje().isoformat()])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "fechadas.xlsx"
        lote = LoteImportacao.objects.create(
            tipo=LoteImportacao.Tipo.TAREFAS, arquivo_nome="f.xlsx", ok=True
        )
        resumo = processar_tarefas(buf, "f.xlsx", lote)
        self.assertEqual(resumo["modo"], "fechadas")
        rel = RelatorioTarefa.objects.get()
        self.assertIsNone(rel.parceiro_id)
        self.assertEqual(rel.total, 2)
        self.assertIn("RJ", rel.mensagem)
        self.assertIn("Rio de Janeiro (RJ)", rel.mensagem)
        self.assertNotIn("(MG) -", rel.mensagem)

    def test_enviar_lote_acumula_resumo(self):
        from unittest.mock import patch

        from gestao.messaging.envio import ResumoEnvio, enviar_tarefas_lote
        from gestao.models import LoteImportacao, RelatorioTarefa
        from gestao.periodo import hoje

        lote = LoteImportacao.objects.create(
            tipo=LoteImportacao.Tipo.TAREFAS, arquivo_nome="t.xlsx", ok=True
        )
        RelatorioTarefa.objects.create(
            lote=lote,
            tipo_relatorio=RelatorioTarefa.TipoRelatorio.ABERTAS,
            parceiro=self.pdv,
            pdv_nome="INOVA MG",
            total=2,
            data_referencia=hoje(),
            mensagem="resumo",
        )
        with patch(
            "gestao.messaging.envio.enviar_tarefa",
            return_value=ResumoEnvio(enviados=1, detalhes=["ok"]),
        ):
            r = enviar_tarefas_lote(lote.id)
        self.assertEqual(r.enviados, 1)
        self.assertTrue(any("INOVA MG" in d for d in r.detalhes))


class VendaIndevidaTests(TestCase):
    def setUp(self):
        self.pdv = Parceiro.objects.create(codigo_pdv="31", nome="INOVA MG")

    def _xlsx_vi(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "BASE_VI"
        ws.append(["ANOMES_ABERTURA", "MOTIVO_CRV", "SUBMOTIVO_CRV", "REDE", "NUMERO_PEDIDO"])
        ws.append(["202601", "INDEVIDA", "DOC", "INOVA MG", "P1"])
        ws.append(["202602", "ERRADA", "END", "INOVA MG", "P2"])
        ws.append(["202601", "INDEVIDA", "DOC", "OUTRO PDV X", "P3"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "vi.xlsx"
        return buf

    def test_processar_vi(self):
        from gestao.models import LoteImportacao, RelatorioVendaIndevida
        from gestao.pipelines.venda_indevida import processar_venda_indevida

        lote = LoteImportacao.objects.create(
            tipo=LoteImportacao.Tipo.VENDA_INDEVIDA, arquivo_nome="vi.xlsx", ok=True
        )
        resumo = processar_venda_indevida(self._xlsx_vi(), "vi.xlsx", lote)
        self.assertEqual(resumo["total_linhas"], 3)
        self.assertEqual(resumo["pdvs"], 2)
        self.assertTrue(RelatorioVendaIndevida.objects.filter(consolidado=True).exists())
        por_pdv = RelatorioVendaIndevida.objects.filter(parceiro=self.pdv, consolidado=False)
        self.assertEqual(por_pdv.count(), 1)
        self.assertEqual(por_pdv.get().total, 2)
        self.assertIn("VENDA INDEVIDA", por_pdv.get().mensagem)


class RecompraTests(TestCase):
    def setUp(self):
        self.pdv = Parceiro.objects.create(codigo_pdv="31", nome="INOVA MG")

    def _xlsx_recompra(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "BASE"
        ws.append(["ds_anomes", "resultado", "REDE", "nr_ordem"])
        ws.append(["202601", "RECOMPRA", "INOVA MG", "1"])
        ws.append(["202602", "SEM RECOMPRA", "INOVA MG", "2"])
        ws.append(["202601", "RECOMPRA", "OUTRO PDV", "3"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "recompra.xlsx"
        return buf

    def test_processar_recompra(self):
        from gestao.models import LoteImportacao, RelatorioRecompra
        from gestao.pipelines.recompra import processar_recompra

        lote = LoteImportacao.objects.create(
            tipo=LoteImportacao.Tipo.RECOMPRA, arquivo_nome="r.xlsx", ok=True
        )
        resumo = processar_recompra(self._xlsx_recompra(), "r.xlsx", lote)
        self.assertEqual(resumo["total_linhas"], 3)
        self.assertEqual(resumo["pdvs"], 2)
        self.assertTrue(RelatorioRecompra.objects.filter(consolidado=True).exists())
        por_pdv = RelatorioRecompra.objects.filter(parceiro=self.pdv, consolidado=False)
        self.assertEqual(por_pdv.count(), 1)
        self.assertEqual(por_pdv.get().total, 2)
        self.assertIn("RECOMPRA", por_pdv.get().mensagem)


@override_settings(
    STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {
            "BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"
        },
    },
)
class GestaoEscopoTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.gestor = User.objects.create_superuser("gesc", "gesc@x.com", "x")
        PerfilStaff.objects.create(user=self.gestor, papel=PerfilStaff.Papel.GESTOR)
        self.spec = User.objects.create_user(
            "specesc", "se@x.com", "x", is_staff=True, first_name="Carla"
        )
        PerfilStaff.objects.create(user=self.spec, papel=PerfilStaff.Papel.ESPECIALISTA)
        self.outro = User.objects.create_user(
            "outroesc", "oe@x.com", "x", is_staff=True, first_name="Diego"
        )
        PerfilStaff.objects.create(user=self.outro, papel=PerfilStaff.Papel.ESPECIALISTA)
        self.pdv_spec = Parceiro.objects.create(
            codigo_pdv="s1", nome="PDV Carla", especialista=self.spec
        )
        self.pdv_outro = Parceiro.objects.create(
            codigo_pdv="s2", nome="PDV Diego", especialista=self.outro
        )

    def test_especialista_meus_e_outros(self):
        self.client.force_login(self.spec)
        meus = self.client.get(reverse("gestao_capilaridade") + "?escopo=meus")
        self.assertContains(meus, "PDV Carla")
        self.assertNotContains(meus, "PDV Diego")
        outros = self.client.get(reverse("gestao_capilaridade") + "?escopo=outros")
        self.assertNotContains(outros, "PDV Diego")
        self.assertNotContains(outros, "PDV Carla")

    def test_gestor_meus_vazio_outros_lista_pdvs(self):
        self.client.force_login(self.gestor)
        meus = self.client.get(reverse("gestao_capilaridade") + "?escopo=meus")
        self.assertNotContains(meus, "PDV Carla")
        self.assertNotContains(meus, "PDV Diego")
        outros = self.client.get(reverse("gestao_capilaridade") + "?escopo=outros")
        self.assertContains(outros, "PDV Carla")
        self.assertContains(outros, "PDV Diego")
        self.assertContains(outros, "Carla")

    def test_especialista_nao_ve_outra_gerencia(self):
        self.spec.perfil_staff.gerencia = "MG INTERIOR"
        self.spec.perfil_staff.save(update_fields=["gerencia"])
        self.outro.perfil_staff.gerencia = "MG METROPOLITANA"
        self.outro.perfil_staff.save(update_fields=["gerencia"])
        self.client.force_login(self.spec)
        outros = self.client.get(reverse("gestao_capilaridade") + "?escopo=outros")
        self.assertNotContains(outros, "PDV Diego")
        meus = self.client.get(reverse("gestao_capilaridade") + "?escopo=meus")
        self.assertContains(meus, "PDV Carla")
        self.assertNotContains(meus, "PDV Diego")

    def test_mesma_gerencia_aparece_em_outros(self):
        self.spec.perfil_staff.gerencia = "MG INTERIOR"
        self.spec.perfil_staff.save(update_fields=["gerencia"])
        self.outro.perfil_staff.gerencia = "MG INTERIOR"
        self.outro.perfil_staff.save(update_fields=["gerencia"])
        self.client.force_login(self.spec)
        outros = self.client.get(reverse("gestao_capilaridade") + "?escopo=outros")
        self.assertContains(outros, "PDV Diego")

    def test_gestor_com_gerencia_so_ve_a_sua(self):
        self.spec.perfil_staff.gerencia = "MG INTERIOR"
        self.spec.perfil_staff.save(update_fields=["gerencia"])
        self.outro.perfil_staff.gerencia = "MG METROPOLITANA"
        self.outro.perfil_staff.save(update_fields=["gerencia"])
        self.gestor.perfil_staff.gerencia = "MG INTERIOR"
        self.gestor.perfil_staff.save(update_fields=["gerencia"])
        self.client.force_login(self.gestor)
        outros = self.client.get(reverse("gestao_capilaridade") + "?escopo=outros")
        self.assertContains(outros, "PDV Carla")
        self.assertNotContains(outros, "PDV Diego")

    def test_admin_troca_gerencia_das_bases(self):
        self.spec.perfil_staff.gerencia = "PP"
        self.spec.perfil_staff.save(update_fields=["gerencia"])
        self.outro.perfil_staff.gerencia = "CMGES"
        self.outro.perfil_staff.save(update_fields=["gerencia"])
        self.gestor.perfil_staff.gerencia = "PP"
        self.gestor.perfil_staff.save(update_fields=["gerencia"])
        self.client.force_login(self.gestor)
        padrao = self.client.get(reverse("gestao_capilaridade") + "?escopo=outros")
        self.assertContains(padrao, "PDV Carla")
        self.assertNotContains(padrao, "PDV Diego")
        r = self.client.post(
            reverse("gestao_gerencia"),
            {
                "gerencia": "CMGES",
                "next": reverse("gestao_capilaridade") + "?escopo=outros",
            },
        )
        self.assertEqual(r.status_code, 302)
        cmges = self.client.get(reverse("gestao_capilaridade") + "?escopo=outros")
        self.assertContains(cmges, "PDV Diego")
        self.assertNotContains(cmges, "PDV Carla")
        self.client.post(
            reverse("gestao_gerencia"),
            {
                "gerencia": "__todas__",
                "next": reverse("gestao_capilaridade") + "?escopo=outros",
            },
        )
        todas = self.client.get(reverse("gestao_capilaridade") + "?escopo=outros")
        self.assertContains(todas, "PDV Carla")
        self.assertContains(todas, "PDV Diego")

    def test_especialista_nao_troca_gerencia(self):
        self.client.force_login(self.spec)
        r = self.client.post(reverse("gestao_gerencia"), {"gerencia": "CMGES"})
        self.assertEqual(r.status_code, 404)

    def test_admin_destinatarios_ve_todas_as_gerencias(self):
        self.spec.perfil_staff.gerencia = "PP"
        self.spec.perfil_staff.save(update_fields=["gerencia"])
        self.outro.perfil_staff.gerencia = "CMGES"
        self.outro.perfil_staff.save(update_fields=["gerencia"])
        self.gestor.perfil_staff.gerencia = "PP"
        self.gestor.perfil_staff.save(update_fields=["gerencia"])
        self.client.force_login(self.gestor)
        r = self.client.get(reverse("gestao_destinatarios"))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "PDV Carla")
        self.assertContains(r, "PDV Diego")
        self.assertContains(r, "CMGES")
        nomes = list(r.context["form"].fields["parceiro"].queryset.values_list("nome", flat=True))
        self.assertEqual(set(nomes), {"PDV Carla", "PDV Diego"})

    def test_admin_parceiros_ve_todas_as_gerencias(self):
        self.spec.perfil_staff.gerencia = "PP"
        self.spec.perfil_staff.save(update_fields=["gerencia"])
        self.outro.perfil_staff.gerencia = "CMGES"
        self.outro.perfil_staff.save(update_fields=["gerencia"])
        self.gestor.perfil_staff.gerencia = "PP"
        self.gestor.perfil_staff.save(update_fields=["gerencia"])
        self.client.force_login(self.gestor)
        r = self.client.get(reverse("parceiros") + "?escopo=outros")
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "PDV Carla")
        self.assertContains(r, "PDV Diego")
        self.assertContains(r, "CMGES")
        self.assertEqual(r.context["gestao_qtd_outros"], 2)

    def test_relatorios_osab_e_tarefas_seguem_a_gerencia(self):
        from gestao.models import HistoricoOSAB, LoteImportacao, RelatorioTarefa
        from gestao.periodo import hoje

        self.spec.perfil_staff.gerencia = "MG INTERIOR"
        self.spec.perfil_staff.save(update_fields=["gerencia"])
        self.outro.perfil_staff.gerencia = "MG METROPOLITANA"
        self.outro.perfil_staff.save(update_fields=["gerencia"])
        HistoricoOSAB.objects.create(
            parceiro=self.pdv_spec,
            descricao_pdv="PDV Carla",
            status="Ok",
            mensagem="Relatório Carla gerência interior",
        )
        HistoricoOSAB.objects.create(
            parceiro=self.pdv_outro,
            descricao_pdv="PDV Diego",
            status="Ok",
            mensagem="Relatório Diego gerência metro",
        )
        lote = LoteImportacao.objects.create(
            tipo=LoteImportacao.Tipo.TAREFAS, arquivo_nome="t.xlsx", ok=True
        )
        RelatorioTarefa.objects.create(
            lote=lote,
            tipo_relatorio=RelatorioTarefa.TipoRelatorio.FECHADAS,
            pdv_nome="MG consolidado",
            data_referencia=hoje(),
            mensagem="Tarefas de todas as gerências",
        )
        RelatorioTarefa.objects.create(
            lote=lote,
            tipo_relatorio=RelatorioTarefa.TipoRelatorio.ABERTAS,
            parceiro=self.pdv_outro,
            pdv_nome="PDV Diego",
            data_referencia=hoje(),
            mensagem="Tarefas só do Diego",
        )
        self.client.force_login(self.spec)
        osab_meus = self.client.get(reverse("gestao_osab") + "?escopo=meus")
        self.assertContains(osab_meus, "Relatório Carla gerência interior")
        self.assertNotContains(osab_meus, "Relatório Diego gerência metro")
        osab_outros = self.client.get(reverse("gestao_osab") + "?escopo=outros")
        self.assertNotContains(osab_outros, "Relatório Diego gerência metro")
        self.assertNotContains(osab_outros, "Relatório Carla gerência interior")
        tarefas = self.client.get(reverse("gestao_tarefas") + "?escopo=outros")
        self.assertNotContains(tarefas, "Tarefas de todas as gerências")
        self.assertNotContains(tarefas, "Tarefas só do Diego")

    def test_fila_nao_muda_com_escopo_gestao(self):
        from tickets.acesso import tickets_visiveis

        self.assertEqual(tickets_visiveis(self.spec).count(), 0)
        from tickets.models import Ticket, TipoDemanda

        Ticket.objects.create(
            parceiro=self.pdv_outro, tipo=TipoDemanda.RESET_SENHA, tt="TTX"
        )
        Ticket.objects.create(
            parceiro=self.pdv_spec, tipo=TipoDemanda.RESET_SENHA, tt="TTY"
        )
        self.assertEqual(tickets_visiveis(self.spec).count(), 1)
        self.assertEqual(tickets_visiveis(self.gestor).count(), 2)


class PlanilhaGestaoTests(TestCase):
    def setUp(self):
        self.pdv = Parceiro.objects.create(codigo_pdv="px", nome="INOVA MG")

    def test_capilaridade_e_osab_geram_xlsx(self):
        from gestao.planilhas import planilha_capilaridade, planilha_osab

        dados, nome = planilha_capilaridade(self.pdv)
        self.assertTrue(nome.endswith(".xlsx"))
        self.assertTrue(dados.startswith(b"PK"))
        dados2, nome2 = planilha_osab(self.pdv)
        self.assertTrue(nome2.endswith(".xlsx"))
        self.assertTrue(dados2.startswith(b"PK"))


@override_settings(
    STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {
            "BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"
        },
    }
)
class ResultadosTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.gestor = User.objects.create_superuser("gres", "gr@x.com", "x")
        PerfilStaff.objects.create(user=self.gestor, papel=PerfilStaff.Papel.GESTOR)
        self.pdv = Parceiro.objects.create(codigo_pdv="r1", nome="INOVA MG")
        self.client.force_login(self.gestor)

    def _dt(self, y, m, d, h=12):
        return timezone.make_aware(datetime(y, m, d, h, 0, 0))

    def test_pagina_resultados(self):
        r = self.client.get(reverse("gestao_resultados"))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "Parcial de vendas")
        self.assertContains(r, "Acumulado do mês")
        self.assertContains(r, "Ranking de VB")
        self.assertContains(r, "Importe a base Excel")
        self.assertContains(r, "Especialistas")
        r2 = self.client.get(reverse("gestao_resultados"), {"aba": "ranking"})
        self.assertContains(r2, "ranking-preview-img")

    def test_parcial_excel_e_imagens(self):
        from gestao.parcial_imagem import imagem_parcial_gerencia, imagem_parcial_pdv, imagem_parcial_especialistas
        from gestao.pipelines.parcial_vendas import (
            agrupar_por_especialista,
            caption_imagem_parcial,
            processar_parcial_excel,
            turno_parcial,
        )

        self.assertEqual(turno_parcial(11), (12, "12h"))
        self.assertEqual(turno_parcial(14), (15, "15h"))
        self.assertEqual(turno_parcial(16), (18, "18h"))
        self.assertEqual(turno_parcial(20), (18, "18h"))

        arquivo = _xlsx(
            [["INOVA MG", 45, 38]],
            ["PDV", "TOTAL", "D-7"],
        )
        resumo = processar_parcial_excel(arquivo, "parcial.xlsx", [self.pdv], turno=15, ano=2026, mes=9)
        self.assertEqual(resumo["total_pp"], 45)
        self.assertEqual(resumo["delta_pp"], 7)
        self.assertEqual(resumo["rotulo_turno"], "15h")
        self.assertEqual(len(resumo["linhas"]), 1)

        png, nome = imagem_parcial_gerencia(resumo)
        self.assertTrue(png.startswith(b"\x89PNG"))
        self.assertIn("Gerencia", nome)

        linha = resumo["linhas"][0]
        png_pdv, _ = imagem_parcial_pdv(linha, resumo)
        self.assertTrue(png_pdv.startswith(b"\x89PNG"))

        grupos = agrupar_por_especialista(resumo["linhas"])
        self.assertEqual(len(grupos), 1)
        png_esp, _ = imagem_parcial_especialistas(resumo, titulo="Carteira PP")
        self.assertTrue(png_esp.startswith(b"\x89PNG"))

        msg = caption_imagem_parcial(resumo, sufixo="Gerência PP")
        self.assertIn("Parcial", msg)

    def test_top_e_piores_sem_repetir(self):
        from gestao.pipelines.parcial_vendas import _top_e_piores

        linhas = [
            {"parceiro_id": 1, "pdv": "A", "vendas": 6, "d7": 0, "delta": 6},
            {"parceiro_id": 2, "pdv": "B", "vendas": 5, "d7": 1, "delta": 4},
            {"parceiro_id": 3, "pdv": "C", "vendas": 2, "d7": 0, "delta": 2},
            {"parceiro_id": 4, "pdv": "VISION", "vendas": 3, "d7": 6, "delta": -3},
        ]
        top, pior = _top_e_piores(linhas)
        self.assertEqual(len(top), 4)
        self.assertEqual(pior, [])

        linhas6 = linhas + [
            {"parceiro_id": 5, "pdv": "E", "vendas": 1, "d7": 0, "delta": 1},
            {"parceiro_id": 6, "pdv": "F", "vendas": 0, "d7": 0, "delta": 0},
        ]
        top6, pior6 = _top_e_piores(linhas6)
        self.assertEqual(len(top6), 5)
        self.assertEqual(len(pior6), 1)
        self.assertEqual(pior6[0]["pdv"], "VISION")
        ids_top = {l["parceiro_id"] for l in top6}
        ids_pior = {l["parceiro_id"] for l in pior6}
        self.assertFalse(ids_top & ids_pior)

    def test_parcial_aplica_escopo_todos(self):
        from gestao.pipelines.parcial_vendas import aplicar_escopo_parcial, montar_parcial

        outro = Parceiro.objects.create(codigo_pdv="r3", nome="OUTRO PDV")
        base = montar_parcial(
            [
                {
                    "parceiro_id": self.pdv.pk,
                    "pdv": self.pdv.nome,
                    "vendas": 10,
                    "d7": 8,
                    "delta": 2,
                    "especialista_id": None,
                    "especialista": "—",
                }
            ],
            ano=2026,
            mes=9,
            turno=15,
            rotulo_turno="15h",
        )
        expandido = aplicar_escopo_parcial(base, [self.pdv, outro])
        self.assertEqual(expandido["qtd_pdvs"], 2)
        self.assertEqual(expandido["total_pp"], 10)
        linha_zero = next(l for l in expandido["linhas"] if l["pdv"] == "OUTRO PDV")
        self.assertEqual(linha_zero["vendas"], 0)

    def test_parcial_completa_ausentes_com_zero(self):
        from gestao.pipelines.parcial_vendas import processar_parcial_excel

        outro = Parceiro.objects.create(codigo_pdv="r2", nome="POINT CELL")
        arquivo = _xlsx(
            [["INOVA MG", 45, 38]],
            ["PDV", "TOTAL", "D-7"],
        )
        resumo = processar_parcial_excel(
            arquivo, "parcial.xlsx", [self.pdv, outro], turno=15, ano=2026, mes=9
        )
        self.assertEqual(resumo["qtd_pdvs"], 2)
        self.assertEqual(resumo["total_pp"], 45)
        linha_zero = next(l for l in resumo["linhas"] if l["pdv"] == "POINT CELL")
        self.assertEqual(linha_zero["vendas"], 0)
        self.assertEqual(linha_zero["d7"], 0)
        self.assertEqual(linha_zero["delta"], 0)

    def test_parcial_import_lê_excel_da_gerencia_inteira(self):
        from gestao.pipelines.parcial_vendas import aplicar_escopo_parcial, processar_parcial_excel

        User = get_user_model()
        esp = User.objects.create_user("esp1", "e@x.com", "x")
        PerfilStaff.objects.create(user=esp, papel=PerfilStaff.Papel.ESPECIALISTA, gerencia="PP")
        PerfilStaff.objects.filter(user=self.gestor).update(gerencia="PP")
        self.pdv.especialista = self.gestor
        self.pdv.save()
        outro = Parceiro.objects.create(codigo_pdv="r9", nome="FABRETTI", especialista=esp)

        arquivo = _xlsx(
            [
                ["INOVA MG", 10, 8],
                ["FABRETTI", 20, 15],
            ],
            ["PDV", "TOTAL", "D-7"],
        )
        todos = list(__import__("tickets.acesso", fromlist=["parceiros_gestao"]).parceiros_gestao(self.gestor, "todos"))
        ids_g = {p.pk for p in todos}
        resumo = processar_parcial_excel(
            arquivo,
            "parcial.xlsx",
            [self.pdv],
            ids_gerencia=ids_g,
            turno=15,
            ano=2026,
            mes=9,
        )
        self.assertIn(str(outro.pk), resumo.get("vendas_por_id", {}))
        expandido = aplicar_escopo_parcial(resumo, todos)
        fabretti = next(l for l in expandido["linhas"] if l["pdv"] == "FABRETTI")
        self.assertEqual(fabretti["vendas"], 20)
        self.assertEqual(expandido["total_pp"], 30)
        self.assertIn("FABRETTI", [l["pdv"] for l in expandido["top5"]])

    def test_agrupar_por_especialista_volume_e_nome_curto(self):
        from gestao.pipelines.parcial_vendas import agrupar_por_especialista, nome_especialista_curto

        self.assertEqual(nome_especialista_curto("CAIO VINICIUS SILVA DE GRAAFF"), "CAIO GRAAFF")
        self.assertEqual(nome_especialista_curto("Rogério Pereira Pacheco"), "Rogério Pacheco")

        linhas = [
            {
                "parceiro_id": 1,
                "pdv": "PDV A",
                "vendas": 5,
                "d7": 0,
                "delta": 5,
                "especialista_id": 1,
                "especialista": "Ana Silva Costa",
            },
            {
                "parceiro_id": 2,
                "pdv": "PDV B",
                "vendas": 10,
                "d7": 0,
                "delta": 10,
                "especialista_id": 2,
                "especialista": "Bruno Santos Lima",
            },
            {
                "parceiro_id": 3,
                "pdv": "PDV C",
                "vendas": 3,
                "d7": 0,
                "delta": 3,
                "especialista_id": 1,
                "especialista": "Ana Silva Costa",
            },
        ]
        grupos = agrupar_por_especialista(linhas)
        self.assertEqual(grupos[0]["especialista"], "Bruno Lima")
        self.assertEqual(grupos[0]["total_vendas"], 10)
        self.assertEqual(grupos[1]["especialista"], "Ana Costa")
        self.assertEqual(grupos[1]["total_vendas"], 8)
        self.assertEqual(grupos[1]["linhas"][0]["vendas"], 5)

    def test_importar_parcial_pela_tela(self):
        arquivo = _xlsx(
            [["INOVA MG", 30, 25]],
            ["PDV", "VENDAS", "D7"],
        )
        r = self.client.post(
            reverse("gestao_resultados"),
            {
                "action": "importar_parcial",
                "escopo": "todos",
                "turno": "15",
                "arquivo": arquivo,
            },
        )
        self.assertEqual(r.status_code, 302)
        self.assertTrue(
            LoteImportacao.objects.filter(tipo=LoteImportacao.Tipo.PARCIAL, ok=True).exists()
        )
        r2 = self.client.get(reverse("gestao_parcial_preview"), {"visao": "gerencia"})
        self.assertEqual(r2.status_code, 200)
        self.assertEqual(r2["Content-Type"], "image/png")

    def test_imagem_ranking_gera_png(self):
        from gestao.pipelines.resultados import montar_ranking
        from gestao.ranking_imagem import imagem_ranking

        ranking = montar_ranking([self.pdv])
        png, nome = imagem_ranking(ranking)
        self.assertTrue(png.startswith(b"\x89PNG"))
        self.assertTrue(nome.endswith(".png"))

    def test_ranking_preview(self):
        r = self.client.get(reverse("gestao_ranking_preview"))
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], "image/png")
        self.assertTrue(r.content.startswith(b"\x89PNG"))

    def test_mensagem_parcial_saudacao_e_frase_do_dia(self):
        from gestao.pipelines.resultados import (
            FRASES_COMERCIAIS,
            caption_parcial_envio,
            frase_comercial_do_dia,
            mensagem_parcial,
            saudacao_horario,
        )

        self.assertEqual(len(set(FRASES_COMERCIAIS)), len(FRASES_COMERCIAIS))
        self.assertGreaterEqual(len(FRASES_COMERCIAIS), 31)
        self.assertEqual(saudacao_horario(8), ("☀️", "Bom dia"))
        self.assertEqual(saudacao_horario(12), ("🌤️", "Boa tarde"))
        self.assertEqual(saudacao_horario(18), ("🌙", "Boa noite"))

        manha = self._dt(2026, 9, 2, 8)
        msg = mensagem_parcial(pdv="INOVA MG", agora=manha, ano=2026, mes=9)
        self.assertIn("Bom dia, *INOVA MG*!", msg)
        self.assertIn("🎯 *Meta do mês — 09/2026*", msg)
        self.assertIn("Segue o acompanhamento. Foco no resultado.", msg)
        self.assertIn(f"_{frase_comercial_do_dia(manha.date())}_", msg)
        self.assertEqual(msg.count("\n\n"), 3)

        tarde = mensagem_parcial(pdv="time", agora=self._dt(2026, 9, 2, 15), ano=2026, mes=9)
        self.assertIn("Boa tarde, *time*!", tarde)
        noite = mensagem_parcial(pdv="time", agora=self._dt(2026, 9, 2, 21), ano=2026, mes=9)
        self.assertIn("Boa noite, *time*!", noite)

        d1, d2 = date(2026, 9, 2), date(2026, 9, 3)
        self.assertNotEqual(frase_comercial_do_dia(d1), frase_comercial_do_dia(d2))
        self.assertEqual(frase_comercial_do_dia(d1), frase_comercial_do_dia(d1))

        gerada = mensagem_parcial(agora=manha, ano=2026, mes=9)
        personalizada = caption_parcial_envio(gerada, self.pdv, agora=manha, ano=2026, mes=9)
        self.assertIn("*INOVA MG*", personalizada)
        self.assertNotIn("*time*", personalizada)
        self.assertEqual(
            caption_parcial_envio("Texto livre", self.pdv),
            "Texto livre",
        )
        vazia = caption_parcial_envio("", self.pdv, agora=manha, ano=2026, mes=9)
        self.assertIn("Bom dia, *INOVA MG*!", vazia)

    def test_import_osab_persiste_municipio(self):
        hoje = timezone.localdate()
        abertura = datetime(hoje.year, hoje.month, 1, 10, 0)
        arquivo = _xlsx(
            [
                [
                    "PMUN",
                    abertura,
                    "TT1",
                    "JOAO",
                    "INOVA MG",
                    abertura,
                    abertura,
                    "Concluído",
                    "500 MEGA",
                    "Betim",
                ]
            ],
            [
                "PEDIDO",
                "DT_REF",
                "MATRICULA_VENDEDOR",
                "NOME_VENDEDOR",
                "DESCRICAO",
                "DATA_ABERTURA",
                "DATA_FECHAMENTO",
                "SITUACAO",
                "VELOCIDADE",
                "MUNICIPIO",
            ],
        )
        processar_osab(arquivo, "OSAB.xlsx", hoje.year, hoje.month)
        self.assertEqual(VendaOSAB.objects.get(pedido="PMUN").municipio, "Betim")

    def test_import_osab_persiste_localidade(self):
        hoje = timezone.localdate()
        abertura = datetime(hoje.year, hoje.month, 1, 10, 0)
        arquivo = _xlsx(
            [
                [
                    "PLoc",
                    abertura,
                    "TT1",
                    "JOAO",
                    "INOVA MG",
                    abertura,
                    abertura,
                    "Concluído",
                    "500 MEGA",
                    "Contagem",
                ]
            ],
            [
                "PEDIDO",
                "DT_REF",
                "MATRICULA_VENDEDOR",
                "NOME_VENDEDOR",
                "DESCRICAO",
                "DATA_ABERTURA",
                "DATA_FECHAMENTO",
                "SITUACAO",
                "VELOCIDADE",
                "LOCALIDADE",
            ],
        )
        processar_osab(arquivo, "OSAB.xlsx", hoje.year, hoje.month)
        self.assertEqual(VendaOSAB.objects.get(pedido="PLoc").municipio, "Contagem")

    def test_acumulado_d0_d1(self):
        from gestao.models import ConfiguracaoOSAB
        from gestao.pipelines.resultados import linhas_acumulado

        ConfiguracaoOSAB.objects.create(
            parceiro=self.pdv, ano=2026, mes=8, meta_vl=10, meta_gross=8
        )
        VendaOSAB.objects.create(
            pedido="A1",
            pdv_nome="INOVA MG",
            parceiro=self.pdv,
            matricula_vendedor="TT1",
            situacao="Concluído",
            data_abertura=self._dt(2026, 8, 20),
            data_fechamento=self._dt(2026, 8, 20),
        )
        VendaOSAB.objects.create(
            pedido="A2",
            pdv_nome="INOVA MG",
            parceiro=self.pdv,
            matricula_vendedor="TT1",
            situacao="Concluído",
            data_abertura=self._dt(2026, 8, 21),
            data_fechamento=self._dt(2026, 8, 21),
        )
        resumo = linhas_acumulado(
            [self.pdv], 2026, 8, data_ref=timezone.localdate().replace(year=2026, month=8, day=26)
        )
        self.assertEqual(resumo["d0"].isoformat(), "2026-08-21")
        self.assertEqual(resumo["d1"].isoformat(), "2026-08-20")
        linha = resumo["linhas"][0]
        self.assertEqual(linha["realizado_vb"], 2)
        self.assertEqual(linha["d0_vb"], 1)
        self.assertEqual(linha["d1_vb"], 1)
        self.assertEqual(linha["pct_vb"], 20.0)

    def test_janela_segunda_feira(self):
        from datetime import date as d

        from gestao.pipelines.resultados import janela_dia_anterior

        self.assertEqual(janela_dia_anterior(d(2026, 9, 7)), (d(2026, 9, 4), d(2026, 9, 6)))
        self.assertEqual(janela_dia_anterior(d(2026, 9, 8)), (d(2026, 9, 7), d(2026, 9, 7)))

    def test_ranking_pontos_e_grupos(self):
        from datetime import date as d

        from gestao.pipelines.resultados import cadastrar_praca_btu, montar_ranking

        cadastrar_praca_btu("Betim")
        self.pdv.data_credenciamento = d(2025, 1, 1)
        self.pdv.save(update_fields=["data_credenciamento"])
        VendaOSAB.objects.create(
            pedido="R1",
            pdv_nome="INOVA MG",
            parceiro=self.pdv,
            matricula_vendedor="TTREG",
            nome_vendedor="ANA",
            situacao="Concluído",
            municipio="Belo Horizonte",
            data_abertura=self._dt(2026, 9, 3),
        )
        VendaOSAB.objects.create(
            pedido="R2",
            pdv_nome="INOVA MG",
            parceiro=self.pdv,
            matricula_vendedor="TTREG",
            nome_vendedor="ANA",
            situacao="Concluído",
            municipio="Betim",
            data_abertura=self._dt(2026, 9, 5),
        )
        VendaOSAB.objects.create(
            pedido="I1",
            pdv_nome="INOVA MG",
            parceiro=self.pdv,
            matricula_vendedor="TTINI",
            nome_vendedor="BIA",
            situacao="Concluído",
            municipio="Contagem",
            data_abertura=self._dt(2026, 9, 6),
        )
        ranking = montar_ranking([self.pdv], data_ref=d(2026, 9, 7))
        self.assertTrue(ranking["periodo"]["oficial"])
        self.assertEqual(ranking["periodo"]["janela_ini"], d(2026, 9, 4))
        regular = ranking["grupos"]["regular"]
        iniciante = ranking["grupos"]["iniciante"]
        self.assertEqual(len(regular), 1)
        self.assertEqual(regular[0]["pontos"], 2.5)
        self.assertEqual(regular[0]["pontos_dia"], 1.5)
        self.assertEqual(regular[0]["vb_btu"], 1)
        self.assertEqual(regular[0]["vb_padrao"], 2)
        self.assertEqual(len(iniciante), 0)

    def test_cadastra_praca_btu(self):
        from gestao.models import PracaBTU

        r = self.client.post(
            reverse("gestao_resultados"),
            {"action": "add_praca_btu", "nome": "Betim"},
        )
        self.assertEqual(r.status_code, 302)
        self.assertTrue(PracaBTU.objects.filter(nome_norm="BETIM").exists())

    def test_gdp_importa_so_especial(self):
        from gestao.models import PracaBTU
        from gestao.pipelines.gdp import processar_gdp

        arquivo = _gdp_xlsx(
            [
                ["MG", "IPATINGA", "NOVO ESPECIAL", "3131307"],
                ["MG", "BETIM", "NOVO REGULAR", "3106705"],
                ["MG", "SABARA", "NOVO ESPECIAL", "3156700"],
                ["PR", "CURITIBA", "PILOTO MVNO", "4106902"],
            ]
        )
        resumo = processar_gdp([(arquivo, "b2c.xlsx")])
        self.assertEqual(resumo["especial_uniao"], 2)
        self.assertEqual(resumo["mg"], 2)
        self.assertTrue(PracaBTU.objects.filter(nome_norm="IPATINGA", ativo=True).exists())
        self.assertTrue(PracaBTU.objects.filter(nome_norm="SABARA", ativo=True).exists())
        self.assertFalse(PracaBTU.objects.filter(nome_norm="BETIM").exists())
        self.assertFalse(PracaBTU.objects.filter(nome_norm="CURITIBA").exists())

    def test_gdp_desativa_quem_saiu_da_oferta(self):
        from gestao.models import PracaBTU
        from gestao.pipelines.gdp import processar_gdp

        processar_gdp(
            [
                (
                    _gdp_xlsx(
                        [
                            ["MG", "IPATINGA", "NOVO ESPECIAL", "1"],
                            ["MG", "LAVRAS", "NOVO ESPECIAL", "2"],
                        ]
                    ),
                    "a.xlsx",
                )
            ]
        )
        processar_gdp(
            [(_gdp_xlsx([["MG", "IPATINGA", "NOVO ESPECIAL", "1"]]), "b.xlsx")]
        )
        self.assertTrue(PracaBTU.objects.get(nome_norm="IPATINGA").ativo)
        self.assertFalse(PracaBTU.objects.get(nome_norm="LAVRAS").ativo)

    def test_gdp_uniao_b2c_b2b(self):
        from gestao.models import PracaBTU
        from gestao.pipelines.gdp import processar_gdp

        b2c = _gdp_xlsx([["MG", "IPATINGA", "NOVO ESPECIAL", "1"]], nome="b2c.xlsx")
        b2b = _gdp_xlsx([["MG", "MURIAE", "NOVO ESPECIAL", "2"]], nome="b2b.xlsx")
        resumo = processar_gdp([(b2c, "b2c.xlsx"), (b2b, "b2b.xlsx")])
        self.assertEqual(resumo["especial_uniao"], 2)
        self.assertEqual(
            set(PracaBTU.objects.filter(ativo=True).values_list("nome_norm", flat=True)),
            {"IPATINGA", "MURIAE"},
        )

    def test_importar_gdp_pela_tela(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        from gestao.models import PracaBTU

        buf = _gdp_xlsx([["MG", "ITABIRA", "NOVO ESPECIAL", "3131703"]])
        upload = SimpleUploadedFile(
            "20260820_B2C_GDP.xlsx",
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        r = self.client.post(
            reverse("gestao_resultados"),
            {"action": "importar_gdp", "arquivo_b2c": upload},
        )
        self.assertEqual(r.status_code, 302)
        self.assertTrue(PracaBTU.objects.filter(nome_norm="ITABIRA", fonte="gdp").exists())

    @skipUnless(
        Path(r"c:\Users\rogge\Downloads\20260820_B2C_GDP.xlsx").exists(),
        "GDP B2C local não encontrado",
    )
    def test_gdp_arquivo_real_b2c(self):
        from gestao.models import PracaBTU
        from gestao.pipelines.gdp import processar_gdp

        caminho = Path(r"c:\Users\rogge\Downloads\20260820_B2C_GDP.xlsx")
        with caminho.open("rb") as fh:
            resumo = processar_gdp([(fh, caminho.name)])
        self.assertEqual(resumo["especial_uniao"], 158)
        self.assertEqual(resumo["mg"], 17)
        self.assertTrue(PracaBTU.objects.filter(nome_norm="IPATINGA", uf="MG", ativo=True).exists())
        self.assertFalse(PracaBTU.objects.filter(nome_norm="BETIM", ativo=True).exists())


def _acompanhamento_xlsx(linhas_meta, linhas_cal, ano_mes=202608):
    wb = Workbook()
    ws = wb.active
    ws.title = "BASE_FISICOS"
    ws.append(["INDB", "INDICADOR", "QTDE", "ANOMES", "SEMANA", "NM_PDV_GRUPO"])
    for linha in linhas_meta:
        ws.append(linha)
    cal = wb.create_sheet("CALENDARIO")
    cal.append(["DATA", "ANOMES", "DU_REG", "DU_GROSS"])
    for linha in linhas_cal:
        cal.append(linha)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "acompanhamento.xlsx"
    return buf


@override_settings(
    STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {
            "BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"
        },
    }
)
class MetasAcompanhamentoTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.gestor = User.objects.create_superuser("gmeta", "gm@x.com", "x")
        PerfilStaff.objects.create(user=self.gestor, papel=PerfilStaff.Papel.GESTOR)
        self.pdv = Parceiro.objects.create(codigo_pdv="m1", nome="INOVA MG")
        self.client.force_login(self.gestor)
        salvar_periodo(2026, 8)

    def test_pagina_tem_importacao(self):
        r = self.client.get(reverse("gestao_configs"))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "Importar acompanhamento semanal")
        self.assertContains(r, "orçado mensal")
        self.assertContains(r, "BASE_FISICOS")
        self.assertContains(r, "Comissão PAP")
        self.assertContains(r, "Super 800 Mb")
        self.assertContains(r, "Ultra 1 Gb com mesh")
        self.assertContains(r, "calendário de pesos")
        self.assertContains(r, "DU VL")

    def test_inputs_du_usam_ponto_nao_virgula(self):
        from gestao.models import ConfiguracaoOSAB
        from gestao.pipelines.calendario import aplicar_nos_pdvs, garantir_mes

        garantir_mes(2026, 8)
        self.pdv.especialista = self.gestor
        self.pdv.save(update_fields=["especialista"])
        ConfiguracaoOSAB.objects.create(
            parceiro=self.pdv, ano=2026, mes=8, meta_vl=80, meta_gross=66
        )
        aplicar_nos_pdvs(2026, 8)
        html = self.client.get(reverse("gestao_configs")).content.decode()
        self.assertIn('name="peso_vl"', html)
        self.assertNotIn('value="0,5"', html)
        self.assertRegex(html, r'name="peso_vl" value="0\.5')
        self.assertRegex(
            html,
            rf'name="p{self.pdv.id}_du_vl"[^>]*value="[0-9]+\.[0-9]',
        )
        self.assertNotRegex(
            html,
            rf'name="p{self.pdv.id}_du_vl"[^>]*value="[0-9]+,[0-9]',
        )

    def test_importa_vl_gross_cap_e_du(self):
        from gestao.models import ConfiguracaoOSAB, MetaCapilaridade
        from gestao.pipelines.metas import processar_metas

        arquivo = _acompanhamento_xlsx(
            [
                ["META", "VL", 113.14, 202608, "S1", "INOVA MG"],
                ["META", "GROSS", 96.71, 202608, "S1", "INOVA MG"],
                ["META", "CAPILARIDADE", 7.77, 202608, "S1", "INOVA MG"],
                ["FCAST", "VL", 107, 202608, None, "INOVA MG"],
                ["FCAST", "GROSS", 84, 202608, None, "INOVA MG"],
                ["FCAST", "CAPILARIDADE", 3, 202608, None, "INOVA MG"],
                ["REAL", "VL", 50, 202608, "S1", "INOVA MG"],
                ["META", "VL", 200, 202608, "S1", "PDV FANTASMA"],
            ],
            [
                [datetime(2026, 8, 1), 202608, 0.4, 0.8],
                [datetime(2026, 8, 3), 202608, 1.0, 1.0],
            ],
        )
        resumo = processar_metas(arquivo, "a.xlsx", 2026, 8)
        self.assertEqual(resumo["atualizados"], 1)
        self.assertEqual(resumo["sem_cadastro_n"], 1)
        self.assertAlmostEqual(resumo["du_vl"], 1.4)
        self.assertAlmostEqual(resumo["du_gross"], 1.8)
        cap = MetaCapilaridade.objects.get(parceiro=self.pdv, ano=2026, mes=8)
        self.assertEqual(cap.meta_vendedores, 8)
        osab = ConfiguracaoOSAB.objects.get(parceiro=self.pdv, ano=2026, mes=8)
        self.assertEqual(osab.meta_vl, 113)
        self.assertEqual(osab.meta_gross, 97)
        self.assertAlmostEqual(osab.du_vl, 1.4)
        self.assertAlmostEqual(osab.du_gross, 1.8)
        pesos = __import__("json").loads(osab.pesos_diarios_vl)
        self.assertAlmostEqual(pesos["1"], 0.4)
        self.assertAlmostEqual(pesos["3"], 1.0)

    def test_preserva_comissao(self):
        from gestao.models import ConfiguracaoOSAB
        from gestao.pipelines.metas import processar_metas

        ConfiguracaoOSAB.objects.create(
            parceiro=self.pdv,
            ano=2026,
            mes=8,
            meta_vl=1,
            comissao_500=15,
            tem_bonus=True,
        )
        arquivo = _acompanhamento_xlsx(
            [["META", "VL", 80, 202608, "S1", "INOVA MG"]],
            [[datetime(2026, 8, 2), 202608, 1, 1]],
        )
        processar_metas(arquivo, "a.xlsx", 2026, 8)
        osab = ConfiguracaoOSAB.objects.get(parceiro=self.pdv, ano=2026, mes=8)
        self.assertEqual(osab.meta_vl, 80)
        self.assertEqual(osab.comissao_500, 15)
        self.assertTrue(osab.tem_bonus)

    def test_importar_pela_tela(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        from gestao.models import ConfiguracaoOSAB

        buf = _acompanhamento_xlsx(
            [["META", "VL", 50, 202608, "S1", "INOVA MG"]],
            [[datetime(2026, 8, 1), 202608, 1, 1]],
        )
        upload = SimpleUploadedFile(
            "BASE_Acompanhamento.xlsx",
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        r = self.client.post(
            reverse("gestao_configs"),
            {"action": "importar_metas", "arquivo": upload},
        )
        self.assertEqual(r.status_code, 302)
        self.assertEqual(
            ConfiguracaoOSAB.objects.get(parceiro=self.pdv, ano=2026, mes=8).meta_vl, 50
        )

    @skipUnless(
        Path(r"c:\Users\rogge\Downloads\BASE_Acompanhamento Semanal_AGO_S4 2 (1).xlsb").exists(),
        "acompanhamento semanal local não encontrado",
    )
    def test_arquivo_real_agosto(self):
        from gestao.models import ConfiguracaoOSAB, MetaCapilaridade
        from gestao.pipelines.metas import processar_metas

        caminho = Path(
            r"c:\Users\rogge\Downloads\BASE_Acompanhamento Semanal_AGO_S4 2 (1).xlsb"
        )
        with caminho.open("rb") as fh:
            resumo = processar_metas(fh, caminho.name, 2026, 8)
        self.assertEqual(resumo["atualizados"], 1)
        osab = ConfiguracaoOSAB.objects.get(parceiro=self.pdv, ano=2026, mes=8)
        # Orçado (INDB=META / coluna ORÇADO), não o plano semanal FCAST 107/84/3.
        self.assertEqual(osab.meta_vl, 113)
        self.assertEqual(osab.meta_gross, 97)
        self.assertAlmostEqual(osab.du_vl, 23.1855, places=3)
        cap = MetaCapilaridade.objects.get(parceiro=self.pdv, ano=2026, mes=8)
        self.assertEqual(cap.meta_vendedores, 8)


class ComissaoPapTests(TestCase):
    def test_classifica_planos_novos(self):
        from gestao.pipelines.comissao import classificar_plano

        self.assertEqual(classificar_plano("400 MEGA"), "400")
        self.assertEqual(classificar_plano("500 MEGA"), "500")
        self.assertEqual(classificar_plano("600 MEGA"), "600")
        self.assertEqual(classificar_plano("700 MEGA"), "800")
        self.assertEqual(classificar_plano("800 MEGA"), "800")
        self.assertEqual(classificar_plano("1000 MEGA"), "1000")
        self.assertEqual(classificar_plano("1 GB"), "1000")
        self.assertEqual(classificar_plano("1000 MEGA", "ULTRA 1GB COM MESH"), "1000_mesh")
        self.assertIsNone(classificar_plano("900 MEGA"))

    def test_receita_regular_btu_e_mesh(self):
        from gestao.models import PoliticaComissao, PracaBTU, VendaOSAB
        from gestao.pipelines.comissao import receita_mix

        pdv = Parceiro.objects.create(codigo_pdv="c1", nome="INOVA MG")
        PoliticaComissao.objects.get_or_create(pk=1)
        PracaBTU.objects.create(nome="Ipatinga", nome_norm="IPATINGA", uf="MG", ativo=True)
        vendas = [
            VendaOSAB(velocidade="500 MEGA", municipio="Belo Horizonte"),
            VendaOSAB(velocidade="700 MEGA", municipio="Contagem"),
            VendaOSAB(velocidade="1000 MEGA", oferta="ULTRA COM MESH", municipio="Betim"),
            VendaOSAB(velocidade="600 MEGA", municipio="Ipatinga"),
            VendaOSAB(velocidade="400 MEGA", municipio="Contagem"),
        ]
        for i, v in enumerate(vendas, start=1):
            v.pedido = f"C{i}"
            v.pdv_nome = pdv.nome
            v.situacao = "Concluído"
        rec = receita_mix(vendas, PoliticaComissao.vigente(), proj_gross=5)
        self.assertEqual(rec["mix"]["500"], 1)
        self.assertEqual(rec["mix"]["800"], 1)
        self.assertEqual(rec["mix"]["1000_mesh"], 1)
        self.assertEqual(rec["mix"]["600"], 1)
        self.assertEqual(rec["mix"]["400"], 1)
        self.assertEqual(rec["mix_btu"], 1)
        self.assertEqual(rec["comissao_realizada"], 350 + 450 + 385 + 245 + 120)

    def test_osab_usa_receita_gerada(self):
        from gestao.models import ConfiguracaoOSAB, HistoricoOSAB, PoliticaComissao

        pdv = Parceiro.objects.create(codigo_pdv="c2", nome="RECORD")
        PoliticaComissao.objects.get_or_create(pk=1)
        hoje = timezone.localdate()
        ConfiguracaoOSAB.objects.create(
            parceiro=pdv,
            ano=hoje.year,
            mes=hoje.month,
            meta_vl=10,
            meta_gross=10,
            du_vl=20,
            du_gross=20,
        )
        abertura = datetime(hoje.year, hoje.month, 1, 10, 0)
        arquivo = _xlsx(
            [
                [
                    "P80",
                    abertura,
                    "TT1",
                    "JOAO",
                    "RECORD",
                    abertura,
                    abertura,
                    "Concluído",
                    "800 MEGA",
                ]
            ],
            [
                "PEDIDO",
                "DT_REF",
                "MATRICULA_VENDEDOR",
                "NOME_VENDEDOR",
                "DESCRICAO",
                "DATA_ABERTURA",
                "DATA_FECHAMENTO",
                "SITUACAO",
                "VELOCIDADE",
            ],
        )
        processar_osab(arquivo, "OSAB.xlsx", hoje.year, hoje.month)
        hist = HistoricoOSAB.objects.get(parceiro=pdv)
        self.assertIn("Relatório financeiro e desempenho", hist.mensagem)
        self.assertIn("Comissão Projetada", hist.mensagem)
        self.assertIn("Mix de Pagamentos", hist.mensagem)
        self.assertIn("GAP Financeiro", hist.mensagem)
        self.assertEqual(hist.detalhes.get("mix_800"), 1)

    def test_salvar_politica_na_tela(self):
        from gestao.models import ConfiguracaoOSAB, PoliticaComissao

        User = get_user_model()
        gestor = User.objects.create_superuser("gpap", "gp@x.com", "x")
        PerfilStaff.objects.create(user=gestor, papel=PerfilStaff.Papel.GESTOR)
        pdv = Parceiro.objects.create(codigo_pdv="c3", nome="VISION")
        ConfiguracaoOSAB.objects.create(parceiro=pdv, ano=2026, mes=8, meta_vl=1)
        salvar_periodo(2026, 8)
        self.client.force_login(gestor)
        r = self.client.post(
            reverse("gestao_configs"),
            {"action": "salvar_politica", "comissao_800": "450", "comissao_500": "350"},
        )
        self.assertEqual(r.status_code, 302)
        self.assertEqual(PoliticaComissao.vigente().comissao_800, 450)
        self.assertEqual(
            ConfiguracaoOSAB.objects.get(parceiro=pdv, ano=2026, mes=8).comissao_700, 450
        )


class CalendarioDuTests(TestCase):
    def test_garante_mes_com_padrao_e_feriado(self):
        from datetime import date as d

        from gestao.models import DiaFiscal
        from gestao.pipelines.calendario import garantir_mes, marcar_feriado, totais_mes

        dias = garantir_mes(2026, 8)
        self.assertEqual(len(dias), 31)
        sab = DiaFiscal.objects.get(data=d(2026, 8, 1))
        self.assertEqual(sab.peso_vl, 0.5)
        self.assertEqual(sab.peso_gross, 0.0)
        marcar_feriado(d(2026, 8, 15), "Padroeira")
        fer = DiaFiscal.objects.get(data=d(2026, 8, 15))
        self.assertTrue(fer.feriado)
        self.assertEqual(fer.peso_vl, 0.0)
        tot = totais_mes(2026, 8)
        self.assertGreater(tot["du_vl"], 0)
        self.assertGreater(tot["du_gross"], 0)

    def test_aplica_du_no_pdv_e_osab_usa_calendario(self):
        from gestao.models import ConfiguracaoOSAB
        from gestao.pipelines.calendario import aplicar_nos_pdvs, garantir_mes

        pdv = Parceiro.objects.create(codigo_pdv="gm", nome="GM TELECOM")
        garantir_mes(2026, 8)
        ConfiguracaoOSAB.objects.create(
            parceiro=pdv, ano=2026, mes=8, meta_vl=100, meta_gross=80
        )
        n = aplicar_nos_pdvs(2026, 8)
        self.assertEqual(n, 1)
        cfg = ConfiguracaoOSAB.objects.get(parceiro=pdv, ano=2026, mes=8)
        self.assertGreater(cfg.du_vl, 0)
        self.assertGreater(cfg.du_gross, 0)
        self.assertTrue(cfg.pesos_diarios_vl)
        from gestao.pipelines.osab import _resolver_pesos

        pesos_vl, _pesos_gr, du_vl, du_gr = _resolver_pesos(cfg, 2026, 8)
        self.assertGreater(du_vl, 0)
        self.assertGreater(du_gr, 0)
        self.assertTrue(pesos_vl)

    def test_import_grava_dia_fiscal(self):
        from datetime import date as d

        from gestao.models import DiaFiscal
        from gestao.pipelines.metas import processar_metas

        pdv = Parceiro.objects.create(codigo_pdv="m2", nome="INOVA MG")
        arquivo = _acompanhamento_xlsx(
            [["META", "VL", 10, 202608, "S1", "INOVA MG"]],
            [[datetime(2026, 8, 3), 202608, 1.0, 1.2]],
        )
        processar_metas(arquivo, "a.xlsx", 2026, 8)
        dia = DiaFiscal.objects.get(data=d(2026, 8, 3))
        self.assertEqual(dia.peso_vl, 1.0)
        self.assertEqual(dia.peso_gross, 1.2)




